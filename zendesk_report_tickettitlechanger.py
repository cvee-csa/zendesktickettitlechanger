#!/usr/bin/env python3
"""
IT Ops Zendesk Tag Report — zendesk_report_tickettitlechanger.py

Pulls all open/pending/on-hold tickets from the three IT Ops Zendesk groups,
applies esc/rarc heuristics, and generates a colour-coded Excel report.
The report is uploaded as a GitHub Actions artifact.

Required environment variables:
    ZENDESK_EMAIL   your Zendesk login email
    ZENDESK_TOKEN   Zendesk API token (Admin > Apps & Integrations > API)
"""

import os, re, time, base64, json, html as _html
from datetime import datetime, timezone, timedelta
from urllib.parse import urlencode

import requests
from openpyxl                import Workbook
from openpyxl.styles         import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils          import get_column_letter

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

# ── Credentials ────────────────────────────────────────────────────────────────
ZENDESK_EMAIL = os.environ["ZENDESK_EMAIL"]
ZENDESK_TOKEN = os.environ["ZENDESK_TOKEN"]
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
DRY_RUN               = os.environ.get("DRY_RUN", "true").lower() == "true"

# ── Automation guardrails ───────────────────────────────────────────────────────────
BOT_REPLY_TAG          = "it_ops_bot_replied"  # Zendesk tag applied after auto-posting
MAX_AUTO_POSTS_PER_RUN = int(os.environ.get("MAX_AUTO_POSTS_PER_RUN", "5"))  # cap per run

# Google Drive upload (optional — set these secrets to enable)
GDRIVE_SA_JSON  = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")  # full JSON key string
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID")            # folder or Shared Drive folder ID

# ── Constants ──────────────────────────────────────────────────────────────────
ZENDESK_DOMAIN = "cloudsecurityalliance.zendesk.com"
BASE_ZD        = f"https://{ZENDESK_DOMAIN}/api/v2"
TICKET_URL     = f"https://{ZENDESK_DOMAIN}/agent/tickets/"
PST            = timezone(timedelta(hours=-8))
_now           = datetime.now(PST)
TODAY          = _now.strftime("%Y-%m-%d")
NOW            = _now.strftime("%Y-%m-%d_%I%M") + ("am" if _now.hour < 12 else "pm")
REPORT_PATH    = f"/tmp/IT_Ops_Tag_Report_{NOW}.xlsx"

IT_OPS_GROUPS = {
    7783360594455:  "IT-Operations",
    37981538647191: "IT-Operations-Projects",
    38675924427287: "IT-Operations-Tasks",
}

IT_OPS_ASSIGNEES = {
    19148954105367: "Neeks",
    5720866160535:  "Jacob",
    38942574549655: "Catherine",
}

RYAN_ID = 396710941733
KURT_ID = 396693552053

# ── SLA thresholds (business hours / days) ────────────────────────────────────
# Source: zendesk-parameters.md — "All thresholds apply to business hours only"
SLA_INITIAL_RESPONSE_HRS = 2    # first IT Ops comment within 2 biz hrs of creation
SLA_REQUESTER_WAIT_HRS   = 4    # max biz hours requester waits for IT Ops reply
SLA_NO_UPDATE_HRS        = 8    # 1 biz day = 8 hrs max since any ticket update
SLA_RESOLUTION_DAYS      = 2    # informational: open > 2 biz days flagged

IT_OPS_AGENT_IDS = {19148954105367, 5720866160535, 38942574549655}  # Neeks, Jacob, Catherine

DEADLINES = {
    "security.txt": datetime(2026, 4, 1, tzinfo=timezone.utc),
}

# ── Google Drive upload ────────────────────────────────────────────────────────
def upload_to_gdrive(file_path):
    """
    Upload file_path to Google Drive (works with both My Drive and Shared Drives).
    Requires GDRIVE_SERVICE_ACCOUNT_JSON and GDRIVE_FOLDER_ID env vars.
    Skips silently if either is missing or google-auth libs are not installed.
    """
    if not GDRIVE_AVAILABLE:
        print("  [Drive] google-auth libraries not installed — skipping upload.")
        return
    if not GDRIVE_SA_JSON or not GDRIVE_FOLDER_ID:
        print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON or GDRIVE_FOLDER_ID not set — skipping.")
        return

    try:
        sa_json = GDRIVE_SA_JSON.strip()
        if not sa_json:
            print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON is blank after stripping whitespace — skipping.")
            return
        creds_info = json.loads(sa_json)

        # Support both service account keys and OAuth user credentials
        if creds_info.get("type") == "service_account":
            creds = service_account.Credentials.from_service_account_info(
                creds_info,
                scopes=["https://www.googleapis.com/auth/drive"],
            )
        else:
            # OAuth user credentials (from get_token.py / InstalledAppFlow)
            from google.oauth2.credentials import Credentials
            creds = Credentials(
                token=creds_info.get("token"),
                refresh_token=creds_info["refresh_token"],
                token_uri=creds_info.get("token_uri", "https://oauth2.googleapis.com/token"),
                client_id=creds_info["client_id"],
                client_secret=creds_info["client_secret"],
                scopes=creds_info.get("scopes"),
            )

        service = build("drive", "v3", credentials=creds)

        file_name = os.path.basename(file_path)
        file_metadata = {"name": file_name, "parents": [GDRIVE_FOLDER_ID]}
        media = MediaFileUpload(
            file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )

        # supportsAllDrives=True makes it work for both Shared Drives and My Drive
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, name, webViewLink",
            supportsAllDrives=True,
        ).execute()

        print(f"  [Drive] Uploaded: {uploaded['name']}")
        print(f"  [Drive] View at : {uploaded.get('webViewLink', '(no link)')}")

    except json.JSONDecodeError as e:
        print(f"  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON is not valid JSON: {e}")
        print(f"  [Drive] Secret starts with: {repr(GDRIVE_SA_JSON[:80])}")
        print("  [Drive] Check that the secret contains the raw JSON (not base64 or a file path).")
    except Exception as e:
        print(f"  [Drive] Upload failed: {e}")


# ── Zendesk API helpers ─────────────────────────────────────────────────────────
def _zd_headers():
    token = base64.b64encode(
        f"{ZENDESK_EMAIL}/token:{ZENDESK_TOKEN}".encode()
    ).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}




def fetch_tickets():
    """
    Fetch IT Ops open/pending/hold tickets using the cursor-based incremental
    export API (/api/v2/incremental/tickets/cursor.json).

    Unlike the Search API, the incremental endpoint:
      - Has no result-count cap
      - Is not subject to per-agent ticket-visibility restrictions
      - Is the same approach used by the proven export_tickets.py script

    We pull all tickets updated in the last LOOKBACK_DAYS days, then filter
    locally to IT Ops groups + open/pending/hold status.
    """
    LOOKBACK_DAYS = 60
    since = int((datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).timestamp())
    print(f"  Using incremental cursor export (last {LOOKBACK_DAYS} days, since={since})")

    url        = f"{BASE_ZD}/incremental/tickets/cursor.json?start_time={since}"
    all_tickets = []
    batch       = 1

    while url:
        r = requests.get(url, headers=_zd_headers(), timeout=60)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 60)))
            continue
        r.raise_for_status()
        data          = r.json()
        batch_tickets = data.get("tickets", [])
        all_tickets.extend(batch_tickets)
        print(f"  Batch {batch}: {len(batch_tickets)} tickets (running total: {len(all_tickets)})")

        if data.get("end_of_stream", False):
            break
        after_url = data.get("after_url")
        if not after_url or after_url == url:
            break
        url    = after_url
        batch += 1
        time.sleep(0.5)

    # Filter locally to IT Ops groups + actionable statuses
    it_ops_ids     = set(IT_OPS_GROUPS.keys())
    target_statuses = {"open", "pending", "hold"}
    tickets = [
        t for t in all_tickets
        if t.get("group_id") in it_ops_ids
        and t.get("status") in target_statuses
    ]

    print(f"\n  Total from API  : {len(all_tickets)} tickets")
    print(f"  After filtering : {len(tickets)} IT Ops open/pending/hold tickets")
    print(f"\n  --- Sample of fetched tickets (first 5) ---")
    for t in tickets[:5]:
        print(f"    #{t['id']} | group={IT_OPS_GROUPS.get(t.get('group_id'), t.get('group_id'))} | "
              f"status={t.get('status')} | subject={t.get('subject','')[:60]}")
    if not tickets:
        print("  WARNING: 0 IT Ops tickets found after filtering.")
        print(f"  (API returned {len(all_tickets)} total tickets in the window)")
    print(f"  ---\n")

    return tickets


def fetch_comments(ticket_id):
    """Return comment list for a ticket. Retries on rate-limit and transient errors."""
    url = f"{BASE_ZD}/tickets/{ticket_id}/comments.json"
    for attempt in range(3):
        r = requests.get(url, headers=_zd_headers(), timeout=30)
        if r.status_code == 429:
            time.sleep(float(r.headers.get("Retry-After", 10)))
            continue
        if r.status_code in (500, 502, 503, 504):
            time.sleep(2 ** attempt)
            continue
        r.raise_for_status()
        return r.json().get("comments", [])
    return []


# ── SLA helpers ────────────────────────────────────────────────────────────────
def _biz_hours_between(start_utc: datetime, end_utc: datetime) -> float:
    """
    Count business hours (Mon–Fri, 09:00–17:00 US/Pacific, UTC-8)
    elapsed between two timezone-aware UTC datetimes.
    Returns a float number of hours (0.0 if end <= start).
    """
    if end_utc <= start_utc:
        return 0.0

    PST_OFFSET = timedelta(hours=-8)   # fixed PST; close enough for SLA flagging
    BIZ_OPEN   = 9
    BIZ_CLOSE  = 17

    # Strip tzinfo and shift to PST so we can do naive date arithmetic
    s = (start_utc + PST_OFFSET).replace(tzinfo=None)
    e = (end_utc   + PST_OFFSET).replace(tzinfo=None)

    total_hours = 0.0
    day = s.replace(hour=0, minute=0, second=0, microsecond=0)
    end_day = e.replace(hour=0, minute=0, second=0, microsecond=0)

    while day <= end_day:
        if day.weekday() < 5:           # Mon=0 … Fri=4
            seg_start = max(s, day.replace(hour=BIZ_OPEN))
            seg_end   = min(e, day.replace(hour=BIZ_CLOSE))
            if seg_end > seg_start:
                total_hours += (seg_end - seg_start).total_seconds() / 3600
        day += timedelta(days=1)

    return total_hours


def _parse_dt(s: str) -> datetime | None:
    """Parse a Zendesk ISO-8601 timestamp string to a timezone-aware datetime."""
    if not s:
        return None
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


def check_sla(ticket: dict, comments: list) -> dict:
    """
    Evaluate SLA compliance for a ticket against the thresholds in
    zendesk-parameters.md.  Returns a dict:
        flags   — list of human-readable flag strings
        display — newline-joined flags, or "OK - Within SLA"
    """
    now        = datetime.now(timezone.utc)
    created_at = _parse_dt(ticket.get("created_at", ""))
    updated_at = _parse_dt(ticket.get("updated_at", ""))
    req_id     = ticket.get("requester_id")

    it_ops_cmts = [c for c in comments if c.get("author_id") in IT_OPS_AGENT_IDS]

    flags = []
    severity = "ok"  # ok, warn, alert

    # 1. Initial response: first IT Ops comment within 2 biz hrs of creation
    if created_at:
        if it_ops_cmts:
            first_resp_dt = _parse_dt(it_ops_cmts[0].get("created_at", ""))
            if first_resp_dt:
                hrs = _biz_hours_between(created_at, first_resp_dt)
                if hrs > SLA_INITIAL_RESPONSE_HRS:
                    flags.append(f"1st resp: {hrs:.0f}h (>{SLA_INITIAL_RESPONSE_HRS}h)")
                    severity = "warn"
        else:
            hrs = _biz_hours_between(created_at, now)
            if hrs > SLA_INITIAL_RESPONSE_HRS:
                flags.append(f"No resp: {hrs:.0f}h")
                severity = "alert"

    # 2. Requester wait: last requester comment unanswered for >4 biz hrs
    if req_id and comments:
        last_req = next((c for c in reversed(comments)
                         if c.get("author_id") == req_id), None)
        if last_req:
            req_dt   = _parse_dt(last_req.get("created_at", ""))
            req_idx  = next((i for i, c in enumerate(comments)
                             if c["id"] == last_req["id"]), -1)
            answered = any(c.get("author_id") in IT_OPS_AGENT_IDS
                           for c in comments[req_idx + 1:])
            if not answered and req_dt:
                hrs = _biz_hours_between(req_dt, now)
                if hrs > SLA_REQUESTER_WAIT_HRS:
                    flags.append(f"Unanswered: {hrs:.0f}h (>{SLA_REQUESTER_WAIT_HRS}h)")
                    severity = "alert"

    # 3. No update: ticket stale for >1 biz day (8 hrs)
    if updated_at:
        hrs = _biz_hours_between(updated_at, now)
        if hrs > SLA_NO_UPDATE_HRS:
            flags.append(f"Stale: {hrs:.0f}h (>{SLA_NO_UPDATE_HRS}h)")
            if severity == "ok":
                severity = "warn"

    # 4. Resolution: informational open-age flag (>2 biz days)
    if created_at:
        open_hrs  = _biz_hours_between(created_at, now)
        open_days = open_hrs / 8
        if open_days > SLA_RESOLUTION_DAYS:
            flags.append(f"Age: {open_days:.0f}d")

    display = " | ".join(flags) if flags else "OK"
    return {
        "flags":    flags,
        "display":  display,
        "severity": severity,  # "ok", "warn", or "alert"
    }


# ── Classification ──────────────────────────────────────────────────────────────
ESC_PATTERNS = [
    r"blocked\s+on",
    r"waiting\s+on\s+(ryan|kurt|dev|r&d|leadership)",
    r"need[s]?\s+(ryan|kurt|dev|leadership|r&d|approval)",
    r"(ryan|kurt)\s+(need[s]?|has\s+to|must|should|is\s+required)",
    r"pending\s+(ryan|kurt|dev|leadership|r&d|approval)",
    r"requires?\s+(ryan|kurt|dev|leadership|approval)",
    r"escalat",
    r"business\s+decision",
    r"leadership\s+decision",
    r"waiting\s+for\s+(a\s+)?(response|decision|approval|review)",
    r"no\s+response\s+(from|since)",
    r"ryan\s+bergsma",
    r"kurt\s+seigfried",
]

RARC_PATTERNS = [
    r"can\s+you\s+confirm",
    r"please\s+confirm",
    r"let\s+me\s+know\s+if",
    r"does\s+this\s+(work|look\s+right|meet)",
    r"is\s+this\s+satisfactory",
    r"can\s+we\s+(go\s+ahead\s+and\s+)?close",
    r"please\s+verify",
    r"good\s+to\s+(go|close)",
    r"everything\s+(look|work|seem)\s+(good|ok|right)",
    r"(ticket|this)\s+can\s+be\s+closed",
    r"let\s+us\s+know\s+when",
    r"confirm.*and\s+(we|i)\s+(will|can|shall)\s+close",
]


# Human-readable descriptions for each pattern
ESC_REASONS = {
    r"blocked\s+on":                                          "Ticket is blocked waiting on someone",
    r"waiting\s+on\s+(ryan|kurt|dev|r&d|leadership)":        "Waiting on Ryan / Kurt / Dev / Leadership",
    r"need[s]?\s+(ryan|kurt|dev|leadership|r&d|approval)":   "Needs input or approval from Ryan / Kurt / Dev / Leadership",
    r"(ryan|kurt)\s+(need[s]?|has\s+to|must|should|is\s+required)": "Action required from Ryan or Kurt",
    r"pending\s+(ryan|kurt|dev|leadership|r&d|approval)":    "Pending action or approval from Ryan / Kurt / Dev / Leadership",
    r"requires?\s+(ryan|kurt|dev|leadership|approval)":      "Requires input or approval",
    r"escalat":                                               "Ticket has been escalated",
    r"business\s+decision":                                  "Awaiting a business decision",
    r"leadership\s+decision":                                "Awaiting a leadership decision",
    r"waiting\s+for\s+(a\s+)?(response|decision|approval|review)": "Waiting for a response, decision, or approval",
    r"no\s+response\s+(from|since)":                         "No response has been received",
    r"ryan\s+bergsma":                                       "Ryan Bergsma is mentioned in the ticket",
    r"kurt\s+seigfried":                                     "Kurt Seigfried is mentioned in the ticket",
}

RARC_REASONS = {
    r"can\s+you\s+confirm":                                  "IT Ops asked the requester to confirm",
    r"please\s+confirm":                                     "IT Ops asked the requester to confirm",
    r"let\s+me\s+know\s+if":                                 "IT Ops is awaiting feedback from the requester",
    r"does\s+this\s+(work|look\s+right|meet)":              "IT Ops asked if the issue has been resolved",
    r"is\s+this\s+satisfactory":                             "IT Ops asked if the resolution is satisfactory",
    r"can\s+we\s+(go\s+ahead\s+and\s+)?close":              "IT Ops asked to close the ticket",
    r"please\s+verify":                                      "IT Ops asked the requester to verify",
    r"good\s+to\s+(go|close)":                               "IT Ops indicated the ticket is ready to close",
    r"everything\s+(look|work|seem)\s+(good|ok|right)":      "IT Ops asked if everything looks good",
    r"(ticket|this)\s+can\s+be\s+closed":                   "IT Ops indicated the ticket can be closed",
    r"let\s+us\s+know\s+when":                               "IT Ops is awaiting confirmation from the requester",
    r"confirm.*and\s+(we|i)\s+(will|can|shall)\s+close":    "IT Ops is waiting to close pending requester confirmation",
}


_EMAIL_RE = re.compile(r"@\S+\.\w{2,}")


def _match_any(patterns, text):
    """Return (pattern, snippet) for the first matching pattern, or (None, None)."""
    # Pre-clean preserving newlines so we can use them as sentence boundaries
    clean = _clean_text(text)
    for p in patterns:
        m = re.search(p, clean, re.IGNORECASE)
        if m:
            before = clean[:m.start()]

            # Walk back to the nearest line / sentence boundary
            best = -1
            for sep in ("\n", ". ", "! ", "? "):
                pos = before.rfind(sep)
                if pos > best:
                    best = pos
            start = (best + 1) if best != -1 else 0

            # If the text between that boundary and the match contains an email
            # address it is almost certainly an email-signature line — skip it
            # and start the snippet from the match itself.
            if _EMAIL_RE.search(clean[start:m.start()]):
                start = m.start()

            # Walk forward to the nearest boundary after the match.
            # Take the minimum (nearest) position across all separator types.
            after_text = clean[m.end():]
            end = len(clean)  # fallback: end of text
            for sep in ("\n", ". ", "! ", "? "):
                pos = after_text.find(sep)
                if pos != -1:
                    candidate = m.end() + pos + len(sep)
                    if candidate < end:
                        end = candidate

            # Collapse newlines to spaces for display — no hard length cap
            snippet = clean[start:end].strip()
            snippet = re.sub(r"\s+", " ", snippet)
            return p, snippet
    return None, None


def classify(ticket, comments):
    """Returns ('esc'|'rarc'|None, reason_str). RARC checked first."""
    status = ticket.get("status", "")

    it_ops_comments = [c for c in comments if c.get("author_id") in IT_OPS_ASSIGNEES]

    if it_ops_comments:
        last_itops     = it_ops_comments[-1]
        last_itops_idx = next(
            (i for i, c in enumerate(comments) if c["id"] == last_itops["id"]), -1
        )
        subsequent        = comments[last_itops_idx + 1:]
        requester_id      = ticket.get("requester_id")
        requester_replied = any(c.get("author_id") == requester_id for c in subsequent)
        matched, snippet  = _match_any(RARC_PATTERNS, last_itops.get("body", ""))
        if matched and not requester_replied:
            description = RARC_REASONS.get(matched, "IT Ops awaiting requester reply")
            return "rarc", f"{description}. No reply from requester yet."

    all_text = " ".join(
        [ticket.get("subject") or "", ticket.get("description") or ""]
        + [c.get("body", "") for c in comments]
    )
    matched, snippet = _match_any(ESC_PATTERNS, all_text)
    if matched:
        description = ESC_REASONS.get(matched, "Blocked on external actor")
        return "esc", description

    if status == "on-hold":
        return "esc", "Ticket is on-hold — pending external action or decision."

    return None, None


def classify_reason_with_claude(client, tag, ticket, comments, regex_reason):
    """
    Use Claude to generate a ticket-specific reason instead of using the
    generic regex-based description. Falls back to regex_reason if no client.
    """
    if not client:
        return regex_reason

    subject = _sanitize_for_api(
        ticket.get("subject") or ticket.get("raw_subject") or "(no title)"
    )[:200]

    last_comments = []
    for c in (comments or [])[-3:]:
        role = "IT Ops" if c.get("author_id") in IT_OPS_ASSIGNEES else "Requester"
        body = _sanitize_for_api(_clean_text(c.get("body") or ""))[:150]
        last_comments.append(f"[{role}] {body}")
    convo = "\n".join(last_comments) if last_comments else "(none)"

    system = (
        "You are an IT operations analyst. Write a single concise sentence "
        "explaining WHY this ticket is classified as {tag}. Be specific to "
        "the actual ticket content — reference the real issue, people involved, "
        "or action needed. Do NOT use generic language like 'awaiting reply'. "
        "Output ONLY the reason sentence, nothing else."
    )
    prompt = (
        f"Tag: {tag.upper()}\n"
        f"Subject: {subject}\n"
        f"Regex classification: {regex_reason}\n"
        f"Recent conversation:\n{convo}\n"
        f"Write a ticket-specific reason."
    )

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=80,
            system=system,
            messages=[{"role": "user", "content": prompt}],
        )
        result = msg.content[0].text.strip()
        return result if result else regex_reason
    except Exception:
        return regex_reason


# ── Ryan escalation SLA ─────────────────────────────────────────────────────────
def ryan_escalation(ticket, comments):
    all_text = " ".join(
        [ticket.get("subject", ""), ticket.get("description", "")]
        + [c.get("body", "") for c in comments]
    ).lower()

    if "ryan bergsma" not in all_text and "ryan" not in all_text:
        return ""

    for keyword, deadline in DEADLINES.items():
        if keyword in all_text:
            days_left = (deadline - datetime.now(timezone.utc)).days
            if days_left <= 14:
                return f"Slack Ryan directly — expires {deadline.strftime('%B %-d')}"

    ryan_mentions = [
        c for c in comments
        if re.search(r"ryan", c.get("body", ""), re.IGNORECASE)
    ]
    if not ryan_mentions:
        return "Tag Ryan in ticket"

    last_dt    = max(
        datetime.fromisoformat(c["created_at"].replace("Z", "+00:00"))
        for c in ryan_mentions
    )
    days_since = (datetime.now(timezone.utc) - last_dt).days

    if days_since < 3:  return "Tag Ryan in ticket"
    if days_since < 7:  return "Slack #internal"
    return "Slack Ryan directly"


# ── Last Ryan tag date ─────────────────────────────────────────────────────
def last_ryan_tag_date(comments):
    """
    Return the date (MM/DD/YYYY) of the most recent comment that mentions
    Ryan Bergsma by name, or an empty string if none.
    """
    ryan_cmts = [
        c for c in comments
        if re.search(r"ryan\s+bergsma", c.get("plain_body") or c.get("body") or "", re.IGNORECASE)
    ]
    if not ryan_cmts:
        return ""
    last_dt = max(
        datetime.fromisoformat(c["created_at"].replace("Z", "+00:00"))
        for c in ryan_cmts
    )
    return last_dt.strftime("%m/%d/%Y")


# ── Automated action description ────────────────────────────────────────────────
def _clean_text(text):
    """Strip HTML entities and markdown noise; preserve newlines for boundary detection."""
    text = _html.unescape(text)                              # &nbsp; → space, &amp; → &, etc.
    text = re.sub(r"\*{1,2}([^*]+)\*{1,2}", r"\1", text)  # **bold** / *italic* → plain
    text = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", text)       # remove markdown images
    text = re.sub(r"[ \t]+", " ", text)     # collapse horizontal whitespace only
    text = re.sub(r"\n[ \t]*\n+", "\n", text)  # collapse multiple blank lines to one
    return text.strip()


def _comment_preview(comment, max_chars=150):
    """Return a short, clean excerpt from a comment body."""
    body = (comment.get("plain_body") or comment.get("body") or "")
    body = _clean_text(body)
    if len(body) > max_chars:
        body = body[:max_chars - 1].rstrip() + "\u2026"
    return body


def already_bot_replied(ticket):
    """Idempotency check: True if the bot has already posted to this ticket."""
    return BOT_REPLY_TAG in (ticket.get("tags") or [])


def post_as_public(tag):
    """ESC tickets must be internal notes (public=False); RARC are public replies."""
    return tag != "esc"


def suggest_reply(client, tag, ticket, comments, action_hint):
    """Ask Claude to draft a ready-to-send Zendesk reply based on ticket context."""
    subject = ticket.get("subject") or ticket.get("raw_subject") or "(no title)"

    convo_lines = []
    for c in (comments or [])[-10:]:
        role       = "IT Ops" if c.get("author_id") in IT_OPS_ASSIGNEES else "Requester"
        visibility = "internal" if not c.get("public", True) else "public"
        date       = (c.get("created_at") or "")[:10]
        body       = _clean_text(c.get("body") or c.get("html_body") or "")[:400]
        convo_lines.append(f"[{date} | {role} | {visibility}]\n{body}")

    convo = "\n\n".join(convo_lines) if convo_lines else "(no conversation yet)"

    if tag == "rarc":
        system = (
            "You are an IT support agent drafting a short, professional Zendesk reply "
            "to the requester. Be specific to this ticket, reference the actual issue, "
            "and use the requester's first name if visible. "
            "Do not use generic filler. Output only the reply text, nothing else."
        )
        prompt = (
            f"Ticket: {subject}\n"
            f"Situation: {action_hint}\n\n"
            f"Recent conversation:\n{convo}\n\n"
            f"Write the public reply to the requester."
        )
    else:  # esc
        system = (
            "You are an IT support agent writing a brief internal Zendesk note for the team. "
            "Include the key ticket context and a specific ask for the person being tagged. "
            "Do not use generic filler. Output only the note text, nothing else."
        )
        prompt = (
            f"Ticket: {subject}\n"
            f"Situation: {action_hint}\n\n"
            f"Recent conversation:\n{convo}\n\n"
            f"Write the internal note."
        )

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            system=system,
            messages=[{"role": "user", "content": prompt}],
        )
        return msg.content[0].text.strip()
    except Exception as e:
        return f"(draft unavailable: {e})"


def _sanitize_for_api(text):
    """Remove null bytes and non-printable characters that cause API 400 errors."""
    if not text:
        return ""
    text = text.replace("\x00", "")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    return text.encode("utf-8", errors="replace").decode("utf-8")


def _rule_based_urgency(sla_flags):
    """Fallback urgency dict when Claude API is unavailable."""
    breach_count = len(sla_flags)
    if breach_count >= 3 or any("No resp" in f or "Unanswered" in f for f in sla_flags):
        level = "HIGH"
    elif breach_count >= 1:
        level = "MEDIUM"
    else:
        level = "LOW"
    return {"level": level, "summary": "", "for_whom": "", "next_step": ""}


def assess_urgency(client, tag, ticket, comments, sla_flags):
    """
    Use Claude to produce a structured urgency assessment for a ticket.
    Returns a dict: {level, summary, for_whom, next_step}
    """
    if not client:
        return _rule_based_urgency(sla_flags)

    subject = _sanitize_for_api(
        ticket.get("subject") or ticket.get("raw_subject") or "(no title)"
    )[:200]
    description = _sanitize_for_api(_clean_text(ticket.get("description") or ""))[:400]

    convo_lines = []
    for c in (comments or [])[-5:]:
        role = "IT Ops" if c.get("author_id") in IT_OPS_ASSIGNEES else "Requester"
        date = (c.get("created_at") or "")[:10]
        body = _sanitize_for_api(_clean_text(c.get("body") or c.get("html_body") or ""))[:200]
        convo_lines.append(f"[{date} | {role}] {body}")

    convo = "\n".join(convo_lines) if convo_lines else "(no conversation)"
    sla_text = "; ".join(sla_flags) if sla_flags else "Within SLA"

    system = (
        "You are an IT operations analyst. Output EXACTLY 4 lines, one per field:\n"
        "URGENCY: HIGH, MEDIUM, or LOW\n"
        "SUMMARY: 1-sentence request summary (be specific to this ticket)\n"
        "FOR: who this is actually for \u2014 a person's name (not always the requester)\n"
        "NEXT: the single most important next action needed, or None\n\n"
        "Guidelines:\n"
        "- esc tickets blocking on leadership \u2192 MEDIUM or HIGH\n"
        "- rarc tickets awaiting confirmation \u2192 LOW or MEDIUM\n"
        "- Multiple SLA breaches \u2192 HIGH\n"
    )
    prompt = (
        f"Tag: {tag.upper()}\nSubject: {subject}\n"
        f"Description: {description}\n"
        f"SLA: {sla_text}\n"
        f"Conversation:\n{convo}\n"
        f"Assess urgency."
    )

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            system=system,
            messages=[{"role": "user", "content": prompt}],
        )
        text = msg.content[0].text.strip()
        # Parse the 4-line response
        result = {"level": "MEDIUM", "summary": "", "for_whom": "", "next_step": ""}
        for line in text.split("\n"):
            line = line.strip()
            if line.upper().startswith("URGENCY:"):
                val = line.split(":", 1)[1].strip().upper()
                result["level"] = val if val in ("HIGH", "MEDIUM", "LOW") else "MEDIUM"
            elif line.upper().startswith("SUMMARY:"):
                result["summary"] = line.split(":", 1)[1].strip()
            elif line.upper().startswith("FOR:"):
                result["for_whom"] = line.split(":", 1)[1].strip()
            elif line.upper().startswith("NEXT:"):
                result["next_step"] = line.split(":", 1)[1].strip()
        return result
    except Exception as e:
        return _rule_based_urgency(sla_flags)


def automated_action(tag, ryan_step, ticket, comments, client=None, dry_run=True, skip_reason=None):
    # ── Gather last internal note and last public reply from IT Ops ──────────
    it_ops_cmts   = [c for c in comments if c.get("author_id") in IT_OPS_ASSIGNEES]
    last_internal = next((c for c in reversed(it_ops_cmts) if not c.get("public", True)), None)
    last_public   = next((c for c in reversed(it_ops_cmts) if c.get("public", True)),  None)

    context_lines = []
    if last_internal:
        date    = (last_internal.get("created_at") or "")[:10]
        author  = IT_OPS_ASSIGNEES.get(last_internal.get("author_id"), "IT Ops")
        preview = _comment_preview(last_internal)
        context_lines.append(f"[Internal note \u2014 {author}, {date}]\n\"{preview}\"")
    if last_public:
        date    = (last_public.get("created_at") or "")[:10]
        author  = IT_OPS_ASSIGNEES.get(last_public.get("author_id"), "IT Ops")
        preview = _comment_preview(last_public)
        context_lines.append(f"[Public reply \u2014 {author}, {date}]\n\"{preview}\"")

    context = "\n".join(context_lines)

    # ── Smart action: ticket-specific next step ──────────────────────────────
    now_utc = datetime.now(timezone.utc)
    actions = []

    if tag == "rarc":
        # How long has the requester been waiting since the last IT Ops public reply?
        if last_public:
            try:
                replied_dt = datetime.fromisoformat(
                    (last_public.get("created_at") or "").replace("Z", "+00:00"))
                wait_hrs = _biz_hours_between(replied_dt, now_utc)
                if wait_hrs >= 24:
                    actions.append(
                        f"Requester silent {wait_hrs:.0f} biz hrs \u2014 "
                        f"consider closing or resolving")
                else:
                    actions.append(
                        f"Follow up with requester \u2014 "
                        f"waiting {wait_hrs:.1f} biz hrs since last IT Ops reply")
            except (ValueError, TypeError):
                actions.append("Send public reply to requester requesting an update")
        else:
            actions.append("No IT Ops public reply found \u2014 send initial reply to requester")

    else:  # esc
        # Ryan-tag age drives the primary action
        ryan_tag_str = last_ryan_tag_date(comments)
        ryan_days    = None
        if ryan_tag_str:
            try:
                ryan_dt   = datetime.strptime(ryan_tag_str, "%m/%d/%Y").replace(tzinfo=timezone.utc)
                ryan_days = (now_utc - ryan_dt).days
            except ValueError:
                pass

        if ryan_days is None:
            actions.append(
                "No Ryan outreach on record \u2014 @mention Ryan in an internal note with a specific ask")
        elif ryan_days == 0:
            actions.append("Ryan tagged today \u2014 allow time to respond")
        elif ryan_days <= 2:
            actions.append(
                f"Ryan tagged {ryan_days}d ago \u2014 allow time to respond")
        elif ryan_days <= 4:
            actions.append(
                f"Follow up with Ryan \u2014 tagged {ryan_days}d ago, no response yet")
        elif ryan_days <= 7:
            actions.append(
                f"Urgent \u2014 Ryan unresponsive {ryan_days}d; post ticket to Slack #internal")

        # Secondary: flag very old open tickets
        created_at = ticket.get("created_at") or ""
        if created_at:
            try:
                created_dt    = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                open_biz_days = _biz_hours_between(created_dt, now_utc) / 8
                if open_biz_days > 10:
                    actions.append(
                        f"Open {open_biz_days:.0f} biz days \u2014 clarify resolution path or close")
            except (ValueError, TypeError):
                pass

    steps = "\n".join(actions)
    base = f"{context}\n\n{steps}" if (context and steps) else (context or steps)

    if client and steps:
        if skip_reason:
            return f"{base}\n\nAuto-reply skipped: {skip_reason}"
        draft = suggest_reply(client, tag, ticket, comments, steps)
        label = "--- DRAFT REPLY [DRY RUN \u2014 not posted] ---" if dry_run else "--- DRAFT REPLY [POSTED TO ZENDESK] ---"
        return f"{base}\n\n{label}\n{draft}"
    return base


# ── Spreadsheet builder ─────────────────────────────────────────────────────────
DARK_HEADER = "1F2D3D"
ESC_FILL    = "FFE8E8";  ALT_ESC    = "FFF0F0"
RARC_FILL   = "E8F4E8";  ALT_RARC   = "F0FAF0"
ESC_BADGE   = "C0392B";  RARC_BADGE = "27AE60"
LINK_COLOR  = "1155CC"

HEADERS = [
    "Tag", "Ticket #", "Group", "Subject", "Days Open",
    "Urgency", "Summary", "For", "Next Step",
    "Reason", "SLA", "Last Updated", "Days Since Ryan",
    "Recommended Action",
]
WIDTHS = [8, 12, 16, 40, 10, 10, 36, 14, 30, 36, 22, 13, 13, 40]

# Shortened group display names
GROUP_SHORT = {
    "IT-Operations":          "IT-Ops",
    "IT-Operations-Projects": "Projects",
    "IT-Operations-Tasks":    "Tasks",
}

SLA_BREACH_BG  = "FFF0F0"   # light red — any breach
SLA_URGENT_BG  = "FFD6D6"   # stronger red — flags
SLA_OK_BG      = "F0FAF0"   # light green — within SLA


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, fc="000000",
          bg=None, wrap=False, align="left", size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fc, size=size)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = _border()
    return c


URGENCY_COLORS = {
    "HIGH":   ("FFCDD2", "B71C1C"),
    "MEDIUM": ("FFF9C4", "F57F17"),
    "LOW":    ("C8E6C9", "1B5E20"),
}
SUMMARY_BG = "E8EAF6"
DORMANT_THRESHOLD = 30  # biz days — tickets older than this are "dormant"
SECTION_BG = "37474F"  # dark separator row for dormant section

# Staleness thresholds (calendar days since last update)
STALE_RECENT  = 7    # green — updated within a week
STALE_AGING   = 14   # yellow — 1-2 weeks
# Anything older → red


def _staleness_colors(updated_str):
    """Return (bg, fc) based on how recently the ticket was updated."""
    try:
        updated_dt = datetime.strptime(updated_str, "%m/%d/%Y")
        age_days   = (datetime.now() - updated_dt).days
    except (ValueError, TypeError):
        return ("FFFFFF", "333333")

    if age_days <= STALE_RECENT:
        return ("C8E6C9", "1B5E20")   # green
    elif age_days <= STALE_AGING:
        return ("FFF9C4", "F57F17")   # yellow
    else:
        return ("FFCDD2", "B71C1C")   # red


def _days_open_colors(days):
    """Return (bg, fc) based on how long the ticket has been open (biz days)."""
    if not isinstance(days, (int, float)):
        return ("FFFFFF", "333333")
    if days <= 5:
        return ("C8E6C9", "1B5E20")   # green
    elif days <= DORMANT_THRESHOLD:
        return ("FFF9C4", "F57F17")   # yellow
    else:
        return ("FFCDD2", "B71C1C")   # red


def _ryan_days_colors(days):
    """Return (bg, fc) based on days since last Ryan contact."""
    if not isinstance(days, (int, float)):
        return ("FFFFFF", "999999")
    if days <= 3:
        return ("C8E6C9", "1B5E20")   # green — recent
    elif days <= 7:
        return ("FFF9C4", "F57F17")   # yellow — aging
    elif days <= 14:
        return ("FFD6D6", "B71C1C")   # light red
    else:
        return ("FFCDD2", "B71C1C")   # red — stale


def build_spreadsheet(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tag Recommendations"

    # ── Count summary rows at the top ──────────────────────────────────────
    esc_n  = sum(1 for r in rows if r["tag"] == "esc")
    rarc_n = sum(1 for r in rows if r["tag"] == "rarc")
    ryan_n = sum(1 for r in rows if isinstance(r.get("ryan_days"), int))

    summary_items = [
        ("Total Tickets:", str(len(rows)), "1F2D3D"),
        ("ESC Tickets:",   str(esc_n),     ESC_BADGE),
        ("RARC Tickets:",  str(rarc_n),    RARC_BADGE),
        ("Ryan Involved:", str(ryan_n),    "BF360C"),
    ]
    for sr, (label, val, color) in enumerate(summary_items, 1):
        c = ws.cell(row=sr, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=12, color=color)
        c.fill = PatternFill("solid", start_color=SUMMARY_BG)
        c.alignment = Alignment(horizontal="right", vertical="center")

        c2 = ws.cell(row=sr, column=2, value=int(val))
        c2.font = Font(name="Arial", bold=True, size=14, color=color)
        c2.fill = PatternFill("solid", start_color=SUMMARY_BG)
        c2.alignment = Alignment(horizontal="left", vertical="center")

        for col in range(3, len(HEADERS) + 1):
            sc = ws.cell(row=sr, column=col)
            sc.fill = PatternFill("solid", start_color=SUMMARY_BG)

    ws.row_dimensions[len(summary_items) + 1].height = 6  # spacer row

    # ── Header row ─────────────────────────────────────────────────────────
    HEADER_ROW = len(summary_items) + 2

    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", start_color=DARK_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HEADER_ROW].height = 24

    # ── Split rows: active vs dormant ──────────────────────────────────────
    active  = [r for r in rows
               if not isinstance(r["days_open"], (int, float))
               or r["days_open"] <= DORMANT_THRESHOLD]
    dormant = [r for r in rows
               if isinstance(r["days_open"], (int, float))
               and r["days_open"] > DORMANT_THRESHOLD]

    cur_row = HEADER_ROW + 1

    def _write_data_row(ws, row_num, r):
        is_esc  = r["tag"] == "esc"
        even    = row_num % 2 == 0
        main_bg = (
            (ESC_FILL if is_esc else RARC_FILL) if not even else
            (ALT_ESC  if is_esc else ALT_RARC)
        )
        badge = ESC_BADGE if is_esc else RARC_BADGE

        # Col 1: Tag
        _cell(ws, row_num, 1, r["tag"].upper(), bold=True, fc=badge, bg=main_bg, align="center")

        # Col 2: Ticket # (hyperlinked)
        url = f"{TICKET_URL}{r['ticket_id']}"
        lnk = ws.cell(row=row_num, column=2, value=r["ticket_id"])
        lnk.font      = Font(name="Arial", bold=True, color=LINK_COLOR, underline="single", size=11)
        lnk.alignment = Alignment(horizontal="center", vertical="top")
        lnk.hyperlink = url
        lnk.fill      = PatternFill("solid", start_color=main_bg)
        lnk.border    = _border()

        # Col 3: Group (shortened)
        short_group = GROUP_SHORT.get(r["group"], r["group"])
        _cell(ws, row_num, 3, short_group, bg=main_bg)

        # Col 4: Subject
        _cell(ws, row_num, 4, r["subject"], bg=main_bg, wrap=True)

        # Col 5: Days Open (colour-coded)
        d_open = r.get("days_open", "")
        do_bg, do_fc = _days_open_colors(d_open)
        _cell(ws, row_num, 5, d_open, fc=do_fc, bg=do_bg, align="center", bold=True)

        # Col 6: Urgency level (colour-coded)
        urgency = r.get("urgency", {})
        urg_level = urgency.get("level", "MEDIUM") if isinstance(urgency, dict) else "MEDIUM"
        urg_bg, urg_fc = URGENCY_COLORS.get(urg_level, (main_bg, "333333"))
        _cell(ws, row_num, 6, urg_level, fc=urg_fc, bg=urg_bg, align="center", bold=True)

        # Col 7: Summary
        summary = urgency.get("summary", "") if isinstance(urgency, dict) else ""
        _cell(ws, row_num, 7, summary, bg=main_bg, wrap=True)

        # Col 8: For
        for_whom = urgency.get("for_whom", "") if isinstance(urgency, dict) else ""
        _cell(ws, row_num, 8, for_whom, bg=main_bg, wrap=True)

        # Col 9: Next Step
        next_step = urgency.get("next_step", "") if isinstance(urgency, dict) else ""
        _cell(ws, row_num, 9, next_step, bg=main_bg, wrap=True)

        # Col 10: Reason (Claude-generated, ticket-specific)
        _cell(ws, row_num, 10, r["reason"], bg=main_bg, wrap=True)

        # Col 11: SLA (condensed, colour-coded by severity)
        sla_text     = r.get("sla_display", "OK")
        sla_severity = r.get("sla_severity", "ok")
        if sla_severity == "alert":
            sla_bg, sla_fc = SLA_URGENT_BG, "8B0000"
        elif sla_severity == "warn":
            sla_bg, sla_fc = SLA_BREACH_BG, "8B0000"
        else:
            sla_bg, sla_fc = SLA_OK_BG, "1B5E20"
        _cell(ws, row_num, 11, sla_text, fc=sla_fc, bg=sla_bg, wrap=True, size=10)

        # Col 12: Last Updated (staleness heat)
        upd = r["last_updated"]
        st_bg, st_fc = _staleness_colors(upd)
        _cell(ws, row_num, 12, upd, fc=st_fc, bg=st_bg, align="center", bold=True)

        # Col 13: Days Since Ryan
        rd = r.get("ryan_days", "")
        rd_bg, rd_fc = _ryan_days_colors(rd)
        display_rd = rd if isinstance(rd, int) else ""
        _cell(ws, row_num, 13, display_rd, fc=rd_fc, bg=rd_bg, align="center", bold=True)

        # Col 14: Recommended Action
        action_text = r.get("action", "")
        _cell(ws, row_num, 14, action_text, bg=main_bg, wrap=True, size=10)

        ws.row_dimensions[row_num].height = 72

    # ── Write active tickets ───────────────────────────────────────────────
    for r in active:
        _write_data_row(ws, cur_row, r)
        cur_row += 1

    # ── Dormant section separator ──────────────────────────────────────────
    if dormant:
        sep_text = f"DORMANT TICKETS \u2014 open > {DORMANT_THRESHOLD} biz days ({len(dormant)} tickets)"
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=len(HEADERS))
        c = ws.cell(row=cur_row, column=1, value=sep_text)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", start_color=SECTION_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        for r in dormant:
            _write_data_row(ws, cur_row, r)
            cur_row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── Executive Summary sheet ───────────────────────────────────────────
    es = wb.create_sheet("Executive Summary", 0)  # insert as first sheet

    # Title
    es.merge_cells("A1:F1")
    title_cell = es.cell(row=1, column=1, value=f"IT Ops Report \u2014 {TODAY}")
    title_cell.font = Font(name="Arial", bold=True, size=16, color="1F2D3D")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    es.row_dimensions[1].height = 32

    # Overview stats
    alert_count = sum(1 for r in rows if r.get("sla_severity") == "alert")
    warn_count  = sum(1 for r in rows if r.get("sla_severity") == "warn")
    high_count  = sum(1 for r in rows
                      if isinstance(r.get("urgency"), dict)
                      and r["urgency"].get("level") == "HIGH")

    stats = [
        ("Total Tickets",    len(rows),  "1F2D3D"),
        ("ESC (Escalated)",  esc_n,      ESC_BADGE),
        ("RARC (Ready Close)", rarc_n,   RARC_BADGE),
        ("HIGH Urgency",     high_count, "B71C1C"),
        ("SLA Alerts",       alert_count, "B71C1C"),
        ("SLA Warnings",     warn_count, "F57F17"),
        ("Ryan Involved",    ryan_n,     "BF360C"),
    ]

    row_num = 3
    for label, val, color in stats:
        es.cell(row=row_num, column=1, value=label).font = Font(
            name="Arial", bold=True, size=11, color="333333")
        v = es.cell(row=row_num, column=2, value=val)
        v.font = Font(name="Arial", bold=True, size=13, color=color)
        v.alignment = Alignment(horizontal="center")
        row_num += 1

    # Top 5 urgent tickets
    row_num += 1
    es.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    sec = es.cell(row=row_num, column=1, value="Top Urgent Tickets")
    sec.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    sec.fill = PatternFill("solid", start_color=DARK_HEADER)
    sec.alignment = Alignment(horizontal="left", vertical="center")
    es.row_dimensions[row_num].height = 24
    row_num += 1

    top_headers = ["#", "Subject", "Urgency", "SLA", "Summary", "Next Step"]
    top_widths  = [10, 40, 10, 22, 36, 30]
    for ci, (h, w) in enumerate(zip(top_headers, top_widths), 1):
        c = es.cell(row=row_num, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="666666")
        c.border = _border()
        es.column_dimensions[get_column_letter(ci)].width = w
    row_num += 1

    # Get top 5 by sort order (already sorted by urgency)
    for r in rows[:5]:
        urg = r.get("urgency", {})
        urg_level = urg.get("level", "") if isinstance(urg, dict) else ""

        tid_cell = es.cell(row=row_num, column=1, value=r["ticket_id"])
        tid_url = f"{TICKET_URL}{r['ticket_id']}"
        tid_cell.font = Font(name="Arial", color=LINK_COLOR, underline="single", size=11)
        tid_cell.hyperlink = tid_url
        tid_cell.border = _border()

        es.cell(row=row_num, column=2, value=r["subject"][:60]).border = _border()

        urg_cell = es.cell(row=row_num, column=3, value=urg_level)
        urg_bg, urg_fc = URGENCY_COLORS.get(urg_level, ("FFFFFF", "333333"))
        urg_cell.font = Font(name="Arial", bold=True, color=urg_fc, size=11)
        urg_cell.fill = PatternFill("solid", start_color=urg_bg)
        urg_cell.alignment = Alignment(horizontal="center")
        urg_cell.border = _border()

        es.cell(row=row_num, column=4, value=r.get("sla_display", "OK")).border = _border()

        summary = urg.get("summary", "") if isinstance(urg, dict) else ""
        es.cell(row=row_num, column=5, value=summary).border = _border()

        next_s = urg.get("next_step", "") if isinstance(urg, dict) else ""
        es.cell(row=row_num, column=6, value=next_s).border = _border()

        es.row_dimensions[row_num].height = 28
        row_num += 1

    wb.save(REPORT_PATH)
    print(f"  Spreadsheet saved \u2192 {REPORT_PATH}")
    return esc_n, rarc_n, ryan_n


# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"IT Ops Tag Report \u2014 {TODAY}")
    print(f"{'='*60}\n")

    # Anthropic client for Claude-generated draft replies (optional)
    client = None
    if ANTHROPIC_AVAILABLE and ANTHROPIC_API_KEY:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        mode = "DRY RUN (not posted)" if DRY_RUN else "LIVE (will post to Zendesk)"
        print(f"  Anthropic client ready \u2014 draft replies enabled [{mode}]")
    else:
        print("  ANTHROPIC_API_KEY not set \u2014 draft replies disabled")

    print("[ 1/3 ] Fetching Zendesk tickets...")
    tickets = fetch_tickets()

    print(f"[ 2/3 ] Analysing {len(tickets)} tickets...")
    rows = []
    skipped = 0
    for idx, ticket in enumerate(tickets, 1):
        tid = ticket["id"]
        print(f"  {idx}/{len(tickets)} \u2014 #{tid}", end="\r")

        comments    = fetch_comments(tid)
        tag, reason = classify(ticket, comments)
        if not tag:
            skipped += 1
            continue

        group_id = ticket.get("group_id")
        group    = IT_OPS_GROUPS.get(group_id, "IT-Operations")
        updated  = (ticket.get("updated_at") or "")[:10]
        try:
            updated = datetime.strptime(updated, "%Y-%m-%d").strftime("%m/%d/%Y")
        except ValueError:
            pass

        last_ryan_tag = last_ryan_tag_date(comments)
        sla           = check_sla(ticket, comments)

        # Compute days since last Ryan contact
        ryan_days = ""
        if last_ryan_tag:
            try:
                ryan_dt = datetime.strptime(last_ryan_tag, "%m/%d/%Y").replace(tzinfo=timezone.utc)
                ryan_days = (datetime.now(timezone.utc) - ryan_dt).days
            except ValueError:
                ryan_days = ""

        # Compute biz days open
        created_at = _parse_dt(ticket.get("created_at", ""))
        days_open = ""
        if created_at:
            days_open = round(_biz_hours_between(created_at, datetime.now(timezone.utc)) / 8, 1)

        # Structured urgency assessment (returns dict with level/summary/for_whom/next_step)
        urgency = assess_urgency(client, tag, ticket, comments, sla["flags"])

        # Claude-generated ticket-specific reason (replaces generic regex description)
        reason = classify_reason_with_claude(client, tag, ticket, comments, reason)

        # Ryan escalation step
        ryan_step = ryan_escalation(ticket, comments)

        # Automated action / recommended next steps + draft reply
        action = automated_action(tag, ryan_step, ticket, comments,
                                  client=client, dry_run=DRY_RUN)

        rows.append({
            "tag":           tag,
            "ticket_id":     tid,
            "group":         group,
            "subject":       ticket.get("subject", ""),
            "urgency":       urgency,
            "reason":        reason,
            "last_updated":  updated,
            "sla_flags":     sla["flags"],
            "sla_display":   sla["display"],
            "sla_severity":  sla["severity"],
            "days_open":     days_open,
            "ryan_days":     ryan_days,
            "action":        action,
        })
        time.sleep(0.15)

    # Sort: urgency level first, then SLA severity, then days open descending
    _urgency_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    _severity_order = {"alert": 0, "warn": 1, "ok": 2}
    def _sort_key(r):
        urg = r.get("urgency", {})
        level = urg.get("level", "MEDIUM") if isinstance(urg, dict) else "MEDIUM"
        urg_rank = _urgency_order.get(level, 1)
        sev_rank = _severity_order.get(r.get("sla_severity", "ok"), 2)
        d_open   = r["days_open"] if isinstance(r["days_open"], (int, float)) else 0
        return (urg_rank, sev_rank, -d_open)
    rows.sort(key=_sort_key)
    esc_count  = sum(1 for r in rows if r["tag"] == "esc")
    rarc_count = sum(1 for r in rows if r["tag"] == "rarc")
    ryan_count = sum(1 for r in rows if isinstance(r.get("ryan_days"), int))
    print(f"\n  {len(rows)} candidates \u2014 {esc_count} esc, {rarc_count} rarc, "
          f"{ryan_count} Ryan-involved "
          f"({skipped} tickets did not match esc/rarc criteria)")

    print("[ 3/3 ] Building spreadsheet...")
    esc_n, rarc_n, ryan_n = build_spreadsheet(rows)
    print(f"  Report: {REPORT_PATH}")

    print("[ + ] Uploading to Google Drive...")
    upload_to_gdrive(REPORT_PATH)

    print(f"\nDone. {len(rows)} tickets reported.\n")


if __name__ == "__main__":
    main()
