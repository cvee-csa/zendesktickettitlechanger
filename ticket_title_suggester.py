"""
Zendesk Ticket Title Suggester (Rule-Based)

Queries Zendesk for open tickets, analyzes their titles using heuristic rules,
and suggests more meaningful titles based on ticket context.

No external AI API required — uses keyword extraction and pattern matching.

Guardrails:
- Rate limiting for Zendesk API calls
- PII redaction in suggestions
- Retry logic with exponential backoff
- Configurable max ticket cap
- Log-only mode by default (no ticket modifications)
- Title length and content validation on suggestions
- Automation/notification email ticket detection
"""

import os
import re
import sys
import json
import time
import logging
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from functools import wraps
from collections import Counter

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

ZENDESK_SUBDOMAIN = (os.environ.get("ZENDESK_SUBDOMAIN") or "").strip()
ZENDESK_EMAIL = (os.environ.get("ZENDESK_EMAIL") or "").strip()
ZENDESK_API_TOKEN = (os.environ.get("ZENDESK_API_TOKEN") or "").strip()

# Google Drive upload (optional)
GDRIVE_SA_JSON   = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID")

ZENDESK_BASE_URL = f"https://{ZENDESK_SUBDOMAIN}.zendesk.com/api/v2"

# How many tickets to process per run
MAX_TICKETS = int(os.environ.get("MAX_TICKETS", "50"))

# Rate limiting: seconds to wait between Zendesk API calls
ZENDESK_RATE_LIMIT_DELAY = float(os.environ.get("ZENDESK_RATE_LIMIT_DELAY", "0.5"))

# Retry configuration
MAX_RETRIES = int(os.environ.get("MAX_RETRIES", "3"))
RETRY_BASE_DELAY = float(os.environ.get("RETRY_BASE_DELAY", "2.0"))

# Maximum allowed title length for suggestions
MAX_TITLE_LENGTH = 150

# Report output — use US/Pacific so DST is handled automatically
_PACIFIC = ZoneInfo("America/Los_Angeles")
_now = datetime.now(_PACIFIC)
NOW = _now.strftime("%Y-%m-%d_%I%M") + ("am" if _now.hour < 12 else "pm")
REPORT_PATH = os.environ.get("OUTPUT_FILE", f"/tmp/IT_Ops_Title_Suggestions_{NOW}.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PII Redaction
# ---------------------------------------------------------------------------

PII_PATTERNS = [
    (re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-.]+)?"), "[EMAIL_REDACTED]"),
    (re.compile(r"\b(\+?1?[-.\\s]?)?\(?\d{3}\)?[-.\\s]?\d{3}[-.\\s]?\d{4}\b"), "[PHONE_REDACTED]"),
    (re.compile(r"\b\d{3}-\d{2}-\d{4}\b"), "[SSN_REDACTED]"),
    (re.compile(r"\b\d{4}[-\\s]?\d{4}[-\\s]?\d{4}[-\\s]?\d{4}\b"), "[CC_REDACTED]"),
    (re.compile(r"\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b"), "[IP_REDACTED]"),
]


def redact_pii(text: str) -> str:
    if not text:
        return text
    for pattern, replacement in PII_PATTERNS:
        text = pattern.sub(replacement, text)
    return text


def title_contains_pii(title: str) -> bool:
    """Check if title contains PII patterns."""
    if not title:
        return False
    for pattern, _ in PII_PATTERNS:
        if pattern.search(title):
            return True
    return False


def strip_html(text: str) -> str:
    """Remove HTML tags and common HTML entities from text."""
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"&nbsp;?", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"&amp;?", " and ", text, flags=re.IGNORECASE)
    text = re.sub(r"&lt;?", "<", text, flags=re.IGNORECASE)
    text = re.sub(r"&gt;?", ">", text, flags=re.IGNORECASE)
    text = re.sub(r"&quot;?", '"', text, flags=re.IGNORECASE)
    text = re.sub(r"&#?\w+;", " ", text)  # any remaining HTML entities
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_subject_line(title: str) -> str:
    """Strip Re:/Fwd: chains, bracketed prefixes, system tags, and org suffixes from subject."""
    if not title:
        return title
    cleaned = re.sub(
        r"^(?:(?:re|fw|fwd)\s*:\s*(?:\[.*?\]\s*)?)+", "", title, flags=re.IGNORECASE
    ).strip()
    cleaned = re.sub(r"\[CloudSecurityAlliance\]\s*", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"\s*[-–—]\s*Cloud Security Alliance\s*$", "", cleaned, flags=re.IGNORECASE).strip()

    # Strip automated system tags like [AAS.ZTAC.ZTResource], [AAS.Tasks.ScheduledTask]
    cleaned = re.sub(r"\s*[-–—]?\s*\[AAS\.[^\]]+\]", "", cleaned, flags=re.IGNORECASE).strip()

    # Treat placeholder values as empty ("null", "N/A", "(no subject)", "none", "untitled")
    if re.match(r"^(?:null|n/?a|none|untitled|\(no\s+subject\)|no\s+subject|—|-|\.+)$", cleaned, re.IGNORECASE):
        cleaned = ""

    return cleaned


# ---------------------------------------------------------------------------
# Spam / marketing detection
# ---------------------------------------------------------------------------

SPAM_PATTERNS = [
    # SEO / guest post spam
    re.compile(r"(seo|guest\s*post|backlink|link\s*building|content\s*collaboration)", re.IGNORECASE),
    re.compile(r"(visitor\s*list|attendee\s*list)\s*(for|of|revealed|uncovered|$)", re.IGNORECASE),
    re.compile(r"\b(visitor|attendee|delegate)\s+list\b", re.IGNORECASE),
    re.compile(r"\b(expo|summit|conference|event).{0,40}\b(list|database|emails?)\b", re.IGNORECASE),
    re.compile(r"(real\s*estate|dormitory|accommodation)\s*(available|opportunity|promotion|sites)", re.IGNORECASE),
    re.compile(r"(high\s*authority|da\s*\d+|domain\s*authority)", re.IGNORECASE),
    re.compile(r"plan\s*your\s*visit\s*to", re.IGNORECASE),
    re.compile(r"(amazing|incredible|exclusive)\s*(content|collaboration|opportunity)", re.IGNORECASE),
    re.compile(r"elevate\s+your\s+.{0,20}\s*seo", re.IGNORECASE),
    re.compile(r"(guest\s*post|sponsored\s*post|paid\s*post)\s*(opportunity|inquiry|proposal)", re.IGNORECASE),
    # Conference attendee / email list vendors (common CSA spam)
    re.compile(r"\d[\d,]*\+?\s*.{0,30}(attendee|professional|executive|contact|delegate)s?.{0,20}(list|database|leads?)", re.IGNORECASE),
    re.compile(r"(attendee|contact|email).{0,20}(list|database|data).{0,20}(conference|summit|expo|event)", re.IGNORECASE),
    re.compile(r"reach\s+(cio|cto|ciso|vp|director|executive)s?\b.{0,40}(conference|summit|expo|event|attending)", re.IGNORECASE),
    re.compile(r"(medical|hospital|physician|healthcare|clinic).{0,30}(email|contact|mailing).{0,20}(list|database|data)", re.IGNORECASE),
    re.compile(r"(b2b|verified|targeted|opt.?in)\s*(email|contact|lead)\s*(list|database|data)", re.IGNORECASE),
    # Job fair / hiring event spam
    re.compile(r"(hiring|recruitment|job)\s*(fair|event|expo)\b", re.IGNORECASE),
    re.compile(r"meet\s+(top\s+)?(hiring|employer|recruiter)s?\b", re.IGNORECASE),
    # Telecom / ISP solicitation
    re.compile(r"(gigabit|broadband|fiber|wireless).{0,20}(service|internet|connectivity).{0,20}(available|arrived|offer|just)", re.IGNORECASE),
    # Generic sales pitch openers
    re.compile(r"^(request\s+for\s+quotation|rfq|sales\s+inquiry|business\s+proposal)\s*$", re.IGNORECASE),
    re.compile(r"opportunity\s+for\s+cloud\s+security\s+alliance", re.IGNORECASE),
]


def is_spam_ticket(title: str, description: str) -> bool:
    """Check if a ticket is spam/marketing based on title and body patterns."""
    text = f"{title} {(description or '')[:500]}"
    for pattern in SPAM_PATTERNS:
        if pattern.search(text):
            return True
    return False


# ---------------------------------------------------------------------------
# Retry logic
# ---------------------------------------------------------------------------


def retry_with_backoff(max_retries: int = MAX_RETRIES, base_delay: float = RETRY_BASE_DELAY):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except requests.RequestException as e:
                    last_exception = e
                    if attempt < max_retries:
                        delay = base_delay * (2 ** attempt)
                        logger.warning(
                            "Attempt %d/%d failed for %s: %s. Retrying in %.1fs...",
                            attempt + 1, max_retries + 1, func.__name__, e, delay,
                        )
                        time.sleep(delay)
                    else:
                        logger.error(
                            "All %d attempts failed for %s: %s",
                            max_retries + 1, func.__name__, e,
                        )
            raise last_exception
        return wrapper
    return decorator


# ---------------------------------------------------------------------------
# Zendesk helpers
# ---------------------------------------------------------------------------


def zendesk_auth():
    return (f"{ZENDESK_EMAIL}/token", ZENDESK_API_TOKEN)


def handle_zendesk_rate_limit(response: requests.Response, attempt: int = 0):
    if response.status_code == 429:
        retry_after = int(response.headers.get("Retry-After", 60))
        logger.warning("Zendesk rate limit hit. Waiting %d seconds...", retry_after)
        time.sleep(retry_after)
        return True
    if response.status_code in (500, 502, 503, 504):
        delay = 2 ** attempt
        logger.warning(
            "Zendesk returned %d. Retrying in %ds (attempt %d)...",
            response.status_code, delay, attempt + 1,
        )
        time.sleep(delay)
        return True
    return False


@retry_with_backoff()
def fetch_open_tickets() -> list[dict]:
    tickets = []
    group = os.environ.get("ZENDESK_GROUP", "IT-Operations")
    query = f'type:ticket status<solved group:"{group}"'
    url = f"{ZENDESK_BASE_URL}/search.json"
    params = {"query": query, "sort_by": "created_at", "sort_order": "desc", "per_page": 100}

    while url and len(tickets) < MAX_TICKETS:
        logger.info("Fetching tickets from: %s", url)
        for attempt in range(MAX_RETRIES + 1):
            resp = requests.get(url, auth=zendesk_auth(), params=params, timeout=30)
            if handle_zendesk_rate_limit(resp, attempt):
                continue
            break

        resp.raise_for_status()
        data = resp.json()
        tickets.extend(data.get("results", []))
        url = data.get("next_page")
        params = None
        time.sleep(ZENDESK_RATE_LIMIT_DELAY)

    return tickets[:MAX_TICKETS]


_user_cache: dict[int, str] = {}


def fetch_user_names(user_ids: list[int]) -> None:
    """Batch-fetch user names from Zendesk and populate _user_cache."""
    unknown = [uid for uid in user_ids if uid and uid not in _user_cache]
    if not unknown:
        return
    # /users/show_many accepts up to 100 IDs per request
    for i in range(0, len(unknown), 100):
        batch = unknown[i:i + 100]
        ids_param = ",".join(str(uid) for uid in batch)
        url = f"{ZENDESK_BASE_URL}/users/show_many.json?ids={ids_param}"
        try:
            resp = requests.get(url, auth=zendesk_auth(), timeout=30)
            resp.raise_for_status()
            for user in resp.json().get("users", []):
                _user_cache[user["id"]] = user.get("name", str(user["id"]))
        except requests.RequestException as e:
            logger.warning("Failed to fetch user names: %s", e)
            for uid in batch:
                _user_cache.setdefault(uid, str(uid))
        time.sleep(ZENDESK_RATE_LIMIT_DELAY)


def get_user_name(user_id: int) -> str:
    """Return cached user name, falling back to the raw ID."""
    return _user_cache.get(user_id, str(user_id))


@retry_with_backoff()
def fetch_ticket_comments(ticket_id: int) -> list[dict]:
    url = f"{ZENDESK_BASE_URL}/tickets/{ticket_id}/comments.json"
    for attempt in range(MAX_RETRIES + 1):
        resp = requests.get(url, auth=zendesk_auth(), params={"per_page": 5}, timeout=30)
        if handle_zendesk_rate_limit(resp, attempt):
            continue
        break

    resp.raise_for_status()
    return resp.json().get("comments", [])


# ---------------------------------------------------------------------------
# Rule-based title analysis
# ---------------------------------------------------------------------------

# Titles that are clearly too vague (case-insensitive exact or near-exact match)
VAGUE_TITLES = {
    "help", "help!", "help me", "please help", "need help",
    "question", "a question", "quick question",
    "issue", "problem", "error", "bug",
    "request", "new request", "support request",
    "urgent", "urgent!", "asap",
    "hi", "hello", "hey", "good morning", "good afternoon",
    "follow up", "follow-up", "following up",
    "update", "status update", "checking in",
    "re:", "fw:", "fwd:",
    "test", "testing", "test ticket",
    "ticket", "new ticket", "support ticket",
    "inquiry", "general inquiry", "membership inquiry",
    "info", "information", "information needed",
    "assistance", "need assistance", "assistance needed",
    "account", "my account", "account issue", "account problem",
    "access", "access issue", "can't access", "cannot access",
    "login", "log in", "login issue", "can't login", "cannot login",
    "password", "password reset", "forgot password",
    "billing", "invoice", "payment",
    "certificate", "certification",
    "registration", "register",
    "question about my account",
    "not working", "doesn't work", "it's not working",
    "broken", "something is broken",
    "none", "n/a", "na", "no subject", "(no subject)",
    "...", ".", "-", "--", "___",
    # Common single-product vague titles
    "csa training", "training", "csa", "star",
    "receipt", "purchase", "order", "transaction",
    "membership", "renewal", "subscription",
}

# Patterns that indicate automated/notification email ticket titles
AUTOMATION_PATTERNS = [
    # Purchase/transaction notifications
    re.compile(r"^purchase\s+notification", re.IGNORECASE),
    re.compile(r"^order\s+(confirmation|receipt|notification)", re.IGNORECASE),
    re.compile(r"^payment\s+(confirmation|receipt|notification|received)", re.IGNORECASE),
    re.compile(r"^transaction\s+(confirmation|receipt|notification)", re.IGNORECASE),
    re.compile(r"^receipt\s+for\s+(your\s+)?purchase", re.IGNORECASE),
    re.compile(r"^your\s+(order|purchase|payment|receipt)", re.IGNORECASE),
    re.compile(r"^invoice\s+(#|number|for)", re.IGNORECASE),
    # Auto-responders and system notifications
    re.compile(r"^(auto[- ]?reply|automatic\s+reply|out\s+of\s+office)", re.IGNORECASE),
    re.compile(r"^(undeliverable|delivery\s+(failure|status)|mail\s+delivery\s+failed)", re.IGNORECASE),
    re.compile(r"^(do\s+not\s+reply|noreply|no-reply)", re.IGNORECASE),
    re.compile(r"^(automated?\s+(message|notification|alert|email|response))", re.IGNORECASE),
    re.compile(r"^(system\s+(notification|alert|message|update))", re.IGNORECASE),
    re.compile(r"^(alert|notification)\s*:", re.IGNORECASE),
    # Subscription/account notifications
    re.compile(r"^(welcome\s+to|thank\s+you\s+for\s+(your\s+)?(order|purchase|registration|signing\s+up))", re.IGNORECASE),
    re.compile(r"^(account\s+(created|activated|confirmation|verification))", re.IGNORECASE),
    re.compile(r"^(password\s+(reset|changed|updated)\s+(request|confirmation|notification))", re.IGNORECASE),
    re.compile(r"^(email\s+(verification|confirmation))", re.IGNORECASE),
    re.compile(r"^(subscription|renewal)\s+(confirmation|notification|reminder)", re.IGNORECASE),
    # Calendar/scheduling
    re.compile(r"^(invitation|invite|accepted|declined|tentative)\s*:", re.IGNORECASE),
    re.compile(r"^(meeting|calendar)\s+(invitation|update|cancellation)", re.IGNORECASE),
    re.compile(r"^(reminder|scheduled)\s*:", re.IGNORECASE),
    # Shipping/delivery
    re.compile(r"^(shipping|delivery|tracking)\s+(confirmation|notification|update)", re.IGNORECASE),
    re.compile(r"^your\s+(package|shipment|order)\s+(has\s+been\s+)?(shipped|delivered)", re.IGNORECASE),
    # Marketing/newsletters
    re.compile(r"^(newsletter|digest|weekly\s+update|monthly\s+update)", re.IGNORECASE),
    re.compile(r"^(special\s+offer|promotion|discount|sale|deal)", re.IGNORECASE),
    # Monitoring/CI/CD
    re.compile(r"^\[?(build|deploy|ci|cd|jenkins|github|gitlab|jira|confluence)\]?\s*", re.IGNORECASE),
    re.compile(r"^(uptime|downtime|monitoring|health\s+check)\s+(alert|notification)", re.IGNORECASE),
    # Contact form submissions
    re.compile(r"^a\s+\w+\s+contact\s+form", re.IGNORECASE),
    re.compile(r"contact\s+form\s+has\s+been\s+(filed|submitted|received)", re.IGNORECASE),
]

# Words/phrases in description that indicate automation-originated tickets
AUTOMATION_BODY_SIGNALS = [
    "this is an automated message",
    "this is an auto-generated",
    "this email was sent automatically",
    "do not reply to this email",
    "do not reply directly",
    "this is a system generated",
    "noreply@", "no-reply@", "donotreply@",
    "automated notification",
    "this message was generated",
    "you are receiving this because",
    "this is a confirmation of your",
    "unsubscribe from these",
    "manage your notification",
    "notification preferences",
    "purchase confirmation",
    "order confirmation",
    "payment receipt",
    "transaction receipt",
    "your receipt from",
    "form has been received from",
    "contact form submission",
]

# Words that don't help identify what the ticket is about
STOP_WORDS = {
    "a", "an", "the", "is", "are", "was", "were", "be", "been", "being",
    "have", "has", "had", "do", "does", "did", "will", "would", "shall",
    "should", "may", "might", "must", "can", "could", "i", "me", "my",
    "we", "our", "you", "your", "he", "she", "it", "they", "them", "their",
    "this", "that", "these", "those", "am", "of", "in", "to", "for", "with",
    "on", "at", "from", "by", "about", "as", "into", "through", "during",
    "before", "after", "above", "below", "between", "out", "off", "over",
    "under", "again", "further", "then", "once", "here", "there", "when",
    "where", "why", "how", "all", "both", "each", "few", "more", "most",
    "other", "some", "such", "no", "nor", "not", "only", "own", "same",
    "so", "than", "too", "very", "just", "because", "but", "and", "or",
    "if", "while", "also", "get", "got", "getting", "need", "needed",
    "want", "wanted", "please", "thanks", "thank", "hi", "hello", "hey",
    "dear", "sir", "madam", "team", "support", "help", "issue", "problem",
    "able", "unable", "trying", "try", "tried", "like", "know", "think",
    "still", "seem", "seems", "new", "using", "use", "used",
    # PII redaction placeholders
    "email_redacted", "phone_redacted", "token_redacted",
    "ip_redacted", "ssn_redacted", "cc_redacted",
    "redacted",
    # HTML artifacts
    "nbsp", "amp", "quot", "lt", "gt", "http", "https", "www", "com",
    "org", "net", "html", "div", "span", "href", "img", "src", "alt",
    "class", "style", "width", "height", "border", "padding", "margin",
    "font", "color", "size", "table", "tbody", "thead", "col",
    # Email boilerplate words
    "sent", "received", "forwarded", "replied", "subject", "date",
    "wrote", "message", "email", "mail",
    # Generic filler
    "really", "actually", "basically", "something", "anything", "everything",
    "nothing", "someone", "anyone", "everyone", "thing", "things",
    "way", "ways", "lot", "lots", "bit", "much", "many",
    "going", "come", "came", "make", "made", "take", "took", "give", "gave",
    "see", "saw", "say", "said", "tell", "told", "ask", "asked",
    "work", "working", "look", "looking", "put", "keep", "let",
    "begin", "began", "seem", "show", "showed",
}

# Words that look like personal names (short, capitalized) — filter from suggestions
# We detect these dynamically rather than maintaining a list

# Known products/features to boost in title suggestions
KNOWN_PRODUCTS = [
    "CSA STAR", "STAR Registry", "STAR Level", "STAR Attestation",
    "CCSK", "CCAK", "CCZT", "Certificate of Cloud Security Knowledge",
    "Certificate of Cloud Auditing Knowledge",
    "Cloud Controls Matrix", "CCM", "CAIQ",
    "AI Safety", "IoT", "Zero Trust",
    "STARWatch", "STAR Watch",
    "GRC Stack", "GRC", "Trusted Cloud Provider",
    "SOC 2", "ISO 27001", "ISO 27017", "ISO 27018",
    "TAISE", "Trusted AI Safety Expert",
]

# Compile product patterns for matching (case-insensitive)
PRODUCT_PATTERNS = [(re.compile(re.escape(p), re.IGNORECASE), p) for p in KNOWN_PRODUCTS]


def is_likely_name(word: str) -> bool:
    """Check if a word looks like a personal name (not a product/acronym)."""
    if not word or len(word) < 2:
        return False
    # All-caps words are acronyms, not names
    if word.isupper() and len(word) <= 6:
        return False
    # Check if it's a known product term
    word_lower = word.lower()
    for _, product in PRODUCT_PATTERNS:
        if word_lower in product.lower().split():
            return False
    # Looks like a name: Capitalized, 3-12 chars, all alpha
    if re.match(r"^[A-Z][a-z]{2,11}$", word):
        return True
    return False


def is_automation_ticket(title: str, description: str) -> bool:
    """Check if a ticket was created by an automation or notification email."""
    # Check title against automation patterns
    for pattern in AUTOMATION_PATTERNS:
        if pattern.search(title.strip()):
            return True

    # Check description for automation signals
    desc_lower = (description or "").lower()[:2000]
    signal_count = sum(1 for signal in AUTOMATION_BODY_SIGNALS if signal in desc_lower)
    # If 2+ signals found in the body, it's very likely automated
    if signal_count >= 2:
        return True

    return False


# ---------------------------------------------------------------------------
# Category detection for suggested titles
# ---------------------------------------------------------------------------

# Authoritative list of valid category prefixes and their descriptions.
# detect_category() must return one of these keys. The descriptions are
# written to a "Category Taxonomy" reference sheet in the XLSX report.
VALID_CATEGORIES = {
    # ── IT Ops core work ────────────────────────────────────────────────────────
    "Access Request":         "Grant or revoke access to systems, tools, groups, or admin roles",
    "Account Issue":          "Locked, expired, or broken accounts; login and authentication failures",
    "Automation":             "Workflow automation, API integrations, webhooks, scripted processes",
    "Configuration":          "Settings changes — Zendesk, email routing, forms, redirects, SLA policies",
    "Data/Reporting":         "Data exports, dashboards, analytics, report generation, lead pulls",
    "Infrastructure":         "DNS, SSL/TLS, servers, cloud hosting, email deliverability, Cloudflare",
    "Onboarding/Offboarding": "New employee provisioning or departing employee deprovisioning",
    "OPS-PROJ":               "Internal IT Ops project tasks (format: [OPS-PROJ | P## | T#/#])",
    "Security":               "Incidents, phishing, credential rotation, vulnerability, PII governance",
    "RIT":                    "CSA's Resource and Information Tracker — access, content, or technical issues",
    "Tooling":                "SaaS platform management, license requests, tool procurement",
    "Working Group":          "WG IT support — calendar, access, page setup, member provisioning",
    # ── IT Ops managed platforms ─────────────────────────────────────────────
    "Skilljar":               "CSA's LMS — course access, enrollment, completion, training issues",
    # ── Routing labels (not IT Ops work — route to appropriate team) ─────────
    "Certification":          "Exam, course, or credential question — route to Education team",
    "Membership":             "Membership inquiry or application — route to Membership team",
    "STAR/Registry":          "STAR program or registry question — route to STAR team",
    # ── Administrative / routing ─────────────────────────────────────────────
    "Automated Notification": "System-generated alerts, vendor notifications, workflow emails",
    "Needs Triage":           "Uncategorized or misdirected — review and route to the right team",
}

# ---------------------------------------------------------------------------
# Two-pass category detection
#
# PASS 1 ("action patterns") — matched against the TITLE only.
#   These detect the *intent* of the ticket (access request, policy/doc work,
#   reporting, etc.) and take priority over tool-name mentions.
#
# PASS 2 ("context patterns") — matched against title + description.
#   These detect the *subject area* (specific program, tool, infra component)
#   and act as a fallback when no action pattern matched.
#
# Order within each pass still matters (first match wins).
# ---------------------------------------------------------------------------

# Pass 1: action-based patterns (title only) — what is the requester DOING?
_ACTION_PATTERNS = [
    # STAR/Registry — very specific program, always wins
    # Notification — automated/scheduled task alerts (match before anything else)
    ("Automated Notification", [
        re.compile(r"^Notification\s*:", re.IGNORECASE),
        re.compile(r"^(?:Alert|Reminder|Scheduled\s+Task)\s*:", re.IGNORECASE),
        # Purchase / registration confirmations
        re.compile(r"^purchase\s+notification\s+for\b", re.IGNORECASE),
        re.compile(r"^registration\s+notification\s+for\b", re.IGNORECASE),
        re.compile(r"^csa\s+support\s+notice\b", re.IGNORECASE),
        # Rippling HR workflow notifications
        re.compile(r"\bemployee\s+status\s+change\b", re.IGNORECASE),
        # Google Workspace / shared drive invitations
        re.compile(r"\byou(?:'ve| have) been added\b", re.IGNORECASE),
        # Vendor "Action Required" product emails (Salesforce, Azure, etc.)
        re.compile(r"^Action\s+Required\s*:", re.IGNORECASE),
        # Generic vendor product/service notice subject lines
        re.compile(r"^(?:Important\s+)?(?:Information|Notice|Update)\s+Regarding\b", re.IGNORECASE),
    ]),
    ("STAR/Registry", [
        re.compile(r"\b(star\s+registry|star\s+level|star\s+attestation|starwatch|caiq|ccm|grc\s+stack|trusted\s+cloud)\b", re.IGNORECASE),
        re.compile(r"\bstar\s+(report|review|submission|listing|entry|profile|contact\s+form)\b", re.IGNORECASE),
        re.compile(r"\ba\s+star\s+contact\s+form\s+has\s+been\s+filed\b", re.IGNORECASE),
        re.compile(r"(registry|listing)\s+.{0,20}(review|update|entry|profile|submission)", re.IGNORECASE),
        re.compile(r"\blogo\s+.{0,15}\bstar\b", re.IGNORECASE),
        re.compile(r"\bvalid.?ai.?ted\b", re.IGNORECASE),
        re.compile(r"\b(aicm|ai\s+controls\s+matrix)\b", re.IGNORECASE),
    ]),
    # Onboarding/Offboarding — new hire provisioning or departing employee deprovisioning
    ("Onboarding/Offboarding", [
        re.compile(r"\b(onboard|offboard)\w*\b", re.IGNORECASE),
        re.compile(r"\bnew\s+(hire|employee|staff|contractor|team\s+member)\b", re.IGNORECASE),
        re.compile(r"\b(account\s+setup|account\s+provision|provision\s+account|provision\s+user)\b", re.IGNORECASE),
        re.compile(r"\b(terminat\w*|deprovision\w*|deactivat\w*)\s+.{0,20}(account|access|user|license)\b", re.IGNORECASE),
        re.compile(r"\b(delete|remov\w*)\s+.{0,30}(user\s+accounts?|m365\s+accounts?|microsoft\s*365?\s+.{0,10}accounts?|google\s+accounts?)\b", re.IGNORECASE),
        re.compile(r"\b(employee|staff|user)\s+(departure|leaving|termination|exit|separation)\b", re.IGNORECASE),
        re.compile(r"\b(last\s+day|first\s+day|start\s+date)\b.{0,30}(account|access|setup|email)\b", re.IGNORECASE),
    ]),
    # Skilljar — CSA's LMS; must be before Billing and Certification so "CSA Tools course access" → Skilljar
    ("Skilljar", [
        re.compile(r"\bskilljar\b", re.IGNORECASE),
        re.compile(r"\bcsa\s+(lms|training\s+platform|learning\s+platform)\b", re.IGNORECASE),
        re.compile(r"\bintro(?:duction)?\s+to\s+csa\s+tools?\b", re.IGNORECASE),
        re.compile(r"\bcsa\s+tools?\s+(course|training|module|lesson|platform)\b", re.IGNORECASE),
    ]),
    # Billing — invoice/payment intent beats certification context (e.g. "Invoice for Exam Bundle")
    ("Billing", [
        re.compile(r"(invoice|receipt|refund|credit|charge|payment|billing|renewal\s+quote|pricing|cost)\b", re.IGNORECASE),
        re.compile(r"(purchase|bought|paid|order)\s+.{0,30}(but|however|issue|problem|wrong|error)", re.IGNORECASE),
        re.compile(r"\bpay\s+.{0,30}(training|course|exam|certification)\b", re.IGNORECASE),
        re.compile(r"\bpayment\s+(option|method|plan)\b", re.IGNORECASE),
    ]),
    # Certification — only exam/course/badge contexts, NOT generic "token" or "credential"
    ("Certification", [
        re.compile(r"\b(ccsk|ccak|cczt|taise|certificate\s+of\s+cloud|trusted\s+ai\s+safety\s+expert)\b", re.IGNORECASE),
        re.compile(r"\b(certification|exam)\b(?!.*(?:stale|rotat|clean))", re.IGNORECASE),
        re.compile(r"\b(badge|voucher)\b", re.IGNORECASE),
        re.compile(r"\bcertificate\s+(not\s+found|missing|lost|expired?|invalid)\b", re.IGNORECASE),
        re.compile(r"(training|course|learning|study)\s+(material|access|platform|portal)", re.IGNORECASE),
    ]),
    # Membership — routing label; wins early so "Membership Inquiry" doesn't fall through
    ("Membership", [
        re.compile(r"\bmembership\s+(inquiry|application|question|request|information|form)\b", re.IGNORECASE),
        re.compile(r"\b(join|apply\s+for)\s+.{0,20}(csa|membership)\b", re.IGNORECASE),
        re.compile(r"\bindividual\s+contributor\s+information\b", re.IGNORECASE),
    ]),
    # RIT — specific CSA tool; must come before Access Request so "RIT access" → RIT, not Access Request
    ("RIT", [
        re.compile(r"\bRIT\b"),
        re.compile(r"\bresource\s+and\s+information\s+tracker\b", re.IGNORECASE),
    ]),
    # Access Request — BEFORE Tooling so "Mailgun access" = Access, not Tooling
    ("Access Request", [
        re.compile(r"(add|remove|grant|revoke|give|need)\s+.{0,30}(access|permission|role|admin|editor|viewer|member)", re.IGNORECASE),
        re.compile(r"\baccess\s+(to|for|on|request)\b", re.IGNORECASE),
        re.compile(r"(add|invite|remove)\s+.{0,20}(user|member|team|group)", re.IGNORECASE),
        re.compile(r"(added?\s+to|removed?\s+from|join)\s+.{0,20}(team|group|channel|org|alias)", re.IGNORECASE),
        re.compile(r"\b(add\w*|be\s+added)\b.{0,40}\balias\b", re.IGNORECASE),
        re.compile(r"\b(sso|oauth|login|sign.?in|password|mfa|2fa)\b", re.IGNORECASE),
        re.compile(r"verify\s+.{0,30}\b(on|in|has)\s+.{0,15}(account|team|access)", re.IGNORECASE),
        re.compile(r"\b(role|privilege|permission)\s+(structure|setup|implement|define|establish)\b", re.IGNORECASE),
        re.compile(r"\b(roles?\s+and\s+(privilege|permission)s?)\b", re.IGNORECASE),
        re.compile(r"\b(audit|review)\s+.{0,15}access\b", re.IGNORECASE),
        re.compile(r"\bpermissions?\s+(to|for)\s+(edit|view|creat|delet|updat|manage|modify)\b", re.IGNORECASE),
        re.compile(r"\b(change|transfer)\s+.{0,40}(owner|ownership)\b", re.IGNORECASE),
    ]),
    # Data/Reporting — "list of users", "run a report", data exports
    ("Data/Reporting", [
        re.compile(r"(dashboard|analytics|metric|data\s+(?:export|import|extract|query|migration))", re.IGNORECASE),
        re.compile(r"(run\s+a?\s*report|generate\s+report|pull\s+report|report\s+of\b)", re.IGNORECASE),
        re.compile(r"\blist\s+of\s+.{0,20}(user|member|account|email|active|staff)", re.IGNORECASE),
        re.compile(r"(data\s+request|data\s+pull|prepkit\s+download)", re.IGNORECASE),
        # Lead/opt-in requests from survey reports
        re.compile(r"\b(opt.?in\s+leads?|lead\s+list|lead\s+pull|survey\s+leads?|report\s+leads?)\b", re.IGNORECASE),
        re.compile(r"\bleads?\s+for\s+.{0,40}(survey|report|program|campaign)\b", re.IGNORECASE),
        re.compile(r"\b(3rd|third).?party.{0,20}leads?\b", re.IGNORECASE),
    ]),
    # Infrastructure (action) — subdomain/domain removal, decommission of hosted resources
    ("Infrastructure", [
        re.compile(r"\b(removal|decommission|retire|shut\s*down)\s+.{0,10}(of\s+)?\w+\.\w+\.\w+", re.IGNORECASE),
    ]),
    # Working Group — WG setup, access, calendar, and page management
    ("Working Group", [
        re.compile(r"\bworking\s+group\b", re.IGNORECASE),
        re.compile(r"\bWG\s+\w", re.IGNORECASE),
        re.compile(r"\bcsa\s+wg\b", re.IGNORECASE),
        re.compile(r"\b(chapter\s+(meeting|page|access|setup|calendar|event|member))\b", re.IGNORECASE),
        re.compile(r"\bmessage\s+for\s+.{0,40}chapter\b", re.IGNORECASE),
    ]),
    # Automation — webhook/API/script automation; BEFORE Configuration to win on "automate workflow"
    ("Automation", [
        re.compile(r"\b(zapier|make\.com|integromat|n8n|automate\.io)\b", re.IGNORECASE),
        re.compile(r"\b(webhook|api\s+integration|api\s+automation|api\s+connect)\b", re.IGNORECASE),
        re.compile(r"\bautomat\w+\s+.{0,30}(workflow|process|task|email|report|trigger)\b", re.IGNORECASE),
        re.compile(r"\b(cron\s+job|scheduled\s+job|batch\s+job|script\s+run)\b", re.IGNORECASE),
        re.compile(r"\b(rpa|robotic\s+process\s+automation)\b", re.IGNORECASE),
    ]),
    # Configuration — BEFORE Documentation so "SLA Policy" = Config, not Documentation
    ("Configuration", [
        re.compile(r"(config|setting|setup|enable|disable|toggle|update|modify|change)\s+.{0,30}(setting|config|option|feature|flag|policy|rule)", re.IGNORECASE),
        re.compile(r"\b(redirect|rewrite|iframe|embed|sitemap|robots\.txt)\b", re.IGNORECASE),
        re.compile(r"(workflow|trigger|schedule|sla\s+polic)", re.IGNORECASE),
        re.compile(r"(set\s*up|creat|build|configur)\w*\s+.{0,30}(form|page|template|landing|portal|widget)", re.IGNORECASE),
        re.compile(r"(link\s+between|connect)\s+.{0,30}(base|table|system|platform)", re.IGNORECASE),
        re.compile(r"\b(convert|migrat)\w*\s+.{0,20}account", re.IGNORECASE),
        re.compile(r"\breview\s+.{0,15}(tenant|infra|m365|office\s*365)", re.IGNORECASE),
        re.compile(r"\b(member\s+)?benefit\s+form\b", re.IGNORECASE),
        re.compile(r"\bcontent\s+group\s+.{0,15}(link|fix|url|update|broken)", re.IGNORECASE),
        re.compile(r"\badd\s+.{0,15}\bto\s+.{0,15}(email|notification|template|sign.?up)", re.IGNORECASE),
        re.compile(r"\bemail\s+forwarding\b", re.IGNORECASE),
        re.compile(r"\b(forward|redirect)\s+.{0,20}(email|mail|inbox)\b", re.IGNORECASE),
    ]),
    # Documentation — policy, publishing, CMS, working group pages, acknowledgements
    ("Documentation", [
        re.compile(r"\b(document|documentation|wiki|guide|readme|runbook|playbook|knowledge\s+base|kb|whitepaper|white\s+paper|publication|glossary)\b", re.IGNORECASE),
        # Content errors in published docs / glossary entries
        re.compile(r"\b(typo|error|mistake|correction|not\s+in\s+agreement|header\s+acronym)\b.{0,40}\b(doc|page|glossary|article|entry|definition)\b", re.IGNORECASE),
        re.compile(r"\b(copies|archived?\s+cop|archived?\s+version).{0,30}(whitepaper|paper|document|publication)\b", re.IGNORECASE),
        re.compile(r"(publish|publishing)\s+.{0,20}(doc|page|article|paper|policy|procedure|process|content)", re.IGNORECASE),
        re.compile(r"\b(policy|procedure|guideline|standard)\s+(document|for|on|about|creation|review|update|ensure|consistency)", re.IGNORECASE),
        re.compile(r"(create|update|review|ensure)\s+.{0,15}(policy|procedure|guideline)", re.IGNORECASE),
        re.compile(r"\bcms\b.{0,40}(?:content|page|site|publish|article|template|creat|manag)|(?:content|page|publish|site|article|template|manag).{0,40}\bcms\b", re.IGNORECASE),
        re.compile(r"(content\s+management|working\s+group\s+page|managing\s+.{0,15}page)", re.IGNORECASE),
        re.compile(r"(acknowledge?ments?|privacy\s+policy|topic\s+filter)\b", re.IGNORECASE),
        re.compile(r"\b(guidance|guideline)\s+(for|on|about)", re.IGNORECASE),
        re.compile(r"\buse\s+guidance\b", re.IGNORECASE),
    ]),
    # Security — tighten: require security-specific context, not just the word "security"
    ("Security", [
        re.compile(r"\b(vulnerability|incident|breach|phishing|malware|ransomware|threat)\b", re.IGNORECASE),
        re.compile(r"\b(security\.txt|pen\s*test|penetration)\b", re.IGNORECASE),
        re.compile(r"\b(pii|data\s+(?:leak|exposure|privacy|protection|flow|governance))\b", re.IGNORECASE),
        re.compile(r"\b(stale\s+token|rotate\s+key|secret\s+rotation|credential\s+rotation)\b", re.IGNORECASE),
        re.compile(r"\bsecurity\s+(audit|review|scan|assessment|incident|alert|patch)\b", re.IGNORECASE),
        re.compile(r"\b(decommission\w*)\s+.{0,30}(pii|data|privacy|governance)", re.IGNORECASE),
        re.compile(r"clean\w*\s+.{0,15}(stale|old|unused)\s+.{0,10}(token|key|secret|credential)", re.IGNORECASE),
        re.compile(r"\b(dpo|data\s+subject\s+request|data\s+deletion|data\s+removal|right\s+to\s+erasure|right\s+to\s+be\s+forgotten)\b", re.IGNORECASE),
        re.compile(r"\b(gdpr|ccpa)\s+.{0,30}(request|deletion|removal|compli)", re.IGNORECASE),
    ]),
]

# Pass 2: context-based patterns (title + description) — what is the subject area?
_CONTEXT_PATTERNS = [
    # Notification — detect automated vendor/system emails by description body signals.
    # Placed first so vendor emails are caught before their product name triggers Tooling.
    ("Automated Notification", [
        re.compile(r"rippling\.com", re.IGNORECASE),
        re.compile(r"\bworkflow\s+triggered\s+for\b", re.IGNORECASE),
        re.compile(r"you\s+received\s+this\s+email\s+because\s+you\s+were\s+(?:invited|subscribed|added)", re.IGNORECASE),
        re.compile(r"addressed\s+to\s+\*cloud\s+security\s+alliance\*", re.IGNORECASE),
        re.compile(r"©\s*\d{4}\s+(?:Rippling|Salesforce|Google LLC|Microsoft)", re.IGNORECASE),
        re.compile(r"\bproduct\s+(?:&|and)\s+service\s+notification\b", re.IGNORECASE),
    ]),
    # STAR/Registry (context fallback) — Valid-AI-ted and STAR submission signals in body
    ("STAR/Registry", [
        re.compile(r"\bvalid.?ai.?ted\b", re.IGNORECASE),
        re.compile(r"\bstar\.watch\b", re.IGNORECASE),
        re.compile(r"\bstar\s+submission\b", re.IGNORECASE),
    ]),
    # Membership (context fallback)
    ("Membership", [
        re.compile(r"\bmembership\s+(inquiry|application|question|request|information)\b", re.IGNORECASE),
        re.compile(r"\bindividual\s+contributor\s+information\b", re.IGNORECASE),
    ]),
    # Skilljar — CSA's LMS; wins before Tooling for training/course issues
    ("Skilljar", [
        re.compile(r"\bskilljar\b", re.IGNORECASE),
        re.compile(r"\bintro(?:duction)?\s+to\s+csa\s+tools?\b", re.IGNORECASE),
        re.compile(r"\bcsa\s+tools?\s+(course|training|module|lesson|platform)\b", re.IGNORECASE),
        re.compile(r"\bcsa\s+(lms|learning\s+platform|training\s+platform)\b", re.IGNORECASE),
    ]),
    # RIT — CSA's Resource and Information Tracker; wins before generic Tooling
    ("RIT", [
        re.compile(r"\bRIT\b"),
        re.compile(r"\bresource\s+and\s+information\s+tracker\b", re.IGNORECASE),
    ]),
    # Tooling BEFORE Infrastructure — so "tableau MCP server" hits Tooling, not Infra
    ("Tooling", [
        re.compile(r"\b(github|gitlab|jira|confluence|slack|zoom|teams|zendesk|airtable|zapier|salesforce|pardot|hubspot|mailgun|surveymonkey)\b", re.IGNORECASE),
        re.compile(r"\b(chatgpt|claude|copilot|anthropic|ai\s+license|ai\s+vendor|qms\s+chat\s*bot)\b", re.IGNORECASE),
        re.compile(r"(consolidat|migrat|decommission|integrat)\w*\s+.{0,30}(tool|platform|service|account|license|subscription)", re.IGNORECASE),
        re.compile(r"\bmcp\s+server\b", re.IGNORECASE),
    ]),
    # Infrastructure — servers, DNS, cloud providers, backups, firmware
    ("Infrastructure", [
        re.compile(r"\b(server|dns|ssl|tls|firewall|vpn|router|firmware)\b", re.IGNORECASE),
        re.compile(r"\bdomain\b(?!\s*\d)", re.IGNORECASE),  # DNS domain, not course "Domain 8"
        re.compile(r"\b(cloudflare|digital.?ocean|aws|azure|gcp)\b", re.IGNORECASE),
        re.compile(r"(dmarc|dkim|spf|mx\s+record|email\s+(?:config|setting|routing|deliverability))", re.IGNORECASE),
        re.compile(r"(deploy|hosting|uptime|outage|downtime|monitoring|backup)\b", re.IGNORECASE),
        re.compile(r"(indexed|crawl|seo)\s+.{0,20}(site|page|url)", re.IGNORECASE),
        re.compile(r"\b(local\s+file|old\s+(?:machine|computer|laptop)|microsoft\s+machine)\b", re.IGNORECASE),
        re.compile(r"\b(malware|severity\s+alert)\b", re.IGNORECASE),
    ]),
    # Billing (fallback if not caught by title)
    ("Billing", [
        re.compile(r"(invoice|receipt|refund|credit|charge|payment|billing|renewal|quote|pricing|cost)\b", re.IGNORECASE),
    ]),
    # Documentation (fallback from description)
    ("Documentation", [
        re.compile(r"(work\s+instruction|knowledge\s+capture|documentation\s+tool|document\s+control)", re.IGNORECASE),
        re.compile(r"(formalize|capture|store)\s+.{0,30}(knowledge|instruction|procedure|process)", re.IGNORECASE),
    ]),
    # Certification (fallback) — only CSA-specific cert terms, not generic "certification"
    ("Certification", [
        re.compile(r"\b(ccsk|ccak|cczt)\b", re.IGNORECASE),
        re.compile(r"\b(exam|badge|voucher)\s+.{0,20}(code|issue|problem|access|not\s+work|missing|expired?)", re.IGNORECASE),
    ]),
    # Account Issue
    ("Account Issue", [
        re.compile(r"(account|profile|user)\s+.{0,20}(expired?|locked|blocked|suspended|disabled|inactive|missing|wrong|invalid)", re.IGNORECASE),
        re.compile(r"(expired?|expir(?:ing|ation))\s+.{0,20}(account|license|subscription|membership|certificate)", re.IGNORECASE),
        re.compile(r"(cannot|can'?t|unable)\s+.{0,20}(log\s*in|sign\s*in|access|authenticate)", re.IGNORECASE),
    ]),
    # Data/Reporting (fallback)
    ("Data/Reporting", [
        re.compile(r"(csv|excel|spreadsheet|tableau|google\s+sheet)", re.IGNORECASE),
        # Lead pull requests in description body
        re.compile(r"\b(get|pull|send|share)\s+.{0,20}leads?\b", re.IGNORECASE),
        re.compile(r"\bsurvey\s+report\b", re.IGNORECASE),
    ]),
    # Configuration (context fallback) — general admin setup detected in description
    ("Configuration", [
        re.compile(r"\bcsa.?admin\b.{0,30}\b(edit|setup|config)\b", re.IGNORECASE),
    ]),
    # Working Group (context fallback) — detected in description
    ("Working Group", [
        re.compile(r"\bworking\s+group\b", re.IGNORECASE),
        re.compile(r"(setting\s+up|standard\s+process\s+for|everything\s+in\s+place\s+for)\s+.{0,50}working\s+group", re.IGNORECASE),
        re.compile(r"\bcsa.?admin\b.{0,30}\b(working.?group)\b", re.IGNORECASE),
    ]),
    # Onboarding/Offboarding (context fallback) — detected in description
    ("Onboarding/Offboarding", [
        re.compile(r"\b(onboard|offboard)\w*\b", re.IGNORECASE),
        re.compile(r"\bnew\s+(hire|employee|staff|contractor)\b", re.IGNORECASE),
        re.compile(r"\b(deprovision|deactivat\w+)\s+.{0,20}(account|user|access)\b", re.IGNORECASE),
        re.compile(r"\b(rippling|bamboohr|workday)\b.{0,50}\b(new|depart|terminat|start)\b", re.IGNORECASE),
    ]),
]


def detect_category(title: str, description: str) -> str:
    """Detect the operational category of a ticket.

    Uses a two-pass approach:
      Pass 1 — action patterns matched against the TITLE only.
               Detects the *intent* (access request, policy work, reporting).
      Pass 2 — context patterns matched against title + description.
               Detects the *subject area* (tool, infra component, program).
    """
    # Pass 1: check title for action-based intent
    for category, patterns in _ACTION_PATTERNS:
        for pat in patterns:
            if pat.search(title):
                if category in VALID_CATEGORIES:
                    return category
                # Category was removed from VALID_CATEGORIES intentionally — fall through
                break

    # Pass 2: check title + description for context/subject
    text = f"{title} {(description or '')[:1500]}"
    for category, patterns in _CONTEXT_PATTERNS:
        for pat in patterns:
            if pat.search(text):
                if category in VALID_CATEGORIES:
                    return category
                # Category was removed from VALID_CATEGORIES intentionally — fall through
                break

    return "Needs Triage"


def _title_from_url(title: str) -> str:
    """Extract a human-readable title from a URL-based ticket title.

    Parses the URL path segments, converts slugs (hyphens/underscores) to
    spaces, drops common web path noise, and returns a Title Cased string.
    Returns empty string if nothing meaningful can be extracted.
    """
    import urllib.parse
    cleaned = title.strip().split("\n")[0].strip()  # take only first line if multiline
    try:
        parsed = urllib.parse.urlparse(cleaned if "://" in cleaned else f"https://{cleaned}")
    except Exception:
        return ""

    path = parsed.path.strip("/")
    if not path:
        # Try query params or fragment
        path = parsed.fragment or parsed.query
    if not path:
        return ""

    # Take the last meaningful path segment
    segments = [s for s in path.split("/") if s and s not in ("index", "index.html", "home", "default")]
    if not segments:
        return ""

    slug = segments[-1]
    # Remove file extensions
    slug = re.sub(r"\.\w{2,5}$", "", slug)
    # Convert slug to words
    words = re.split(r"[-_+]+", slug)
    words = [w for w in words if w and len(w) > 1]
    if not words:
        return ""

    readable = " ".join(words)
    readable = normalize_title_grammar(readable)

    # Reject if result is too short or still looks like noise
    if len(readable) < 5 or len(readable.split()) < 2:
        return ""

    return readable


def is_url_title(title: str) -> bool:
    """Check if a title is primarily a URL or URL fragment."""
    cleaned = title.strip()
    # Title starts with a URL (i.e. the URL *is* the title)
    if re.match(r"^https?://", cleaned, re.IGNORECASE):
        return True
    if re.match(r"^www\.", cleaned, re.IGNORECASE):
        return True
    # Title contains a URL — only flag if the non-URL text is too short to be meaningful
    url_match = re.search(r"https?://\S+", cleaned)
    if url_match:
        non_url_text = cleaned[:url_match.start()] + cleaned[url_match.end():]
        non_url_words = [w for w in non_url_text.split() if len(w) > 2]
        if len(non_url_words) < 3:
            return True
    return False


def is_vague_title(title: str) -> bool:
    """Check if a title is too vague/generic to be useful."""
    cleaned = title.strip().lower()
    cleaned = re.sub(r"^(re|fw|fwd)\s*:\s*", "", cleaned).strip()

    # URL-based titles are always vague
    if is_url_title(title):
        return True

    # Exact match against vague titles
    if cleaned in VAGUE_TITLES:
        return True

    # Too short (under 3 words or under 10 chars)
    words = cleaned.split()
    if len(words) < 3 or len(cleaned) < 10:
        return True

    # Only contains stop words
    meaningful_words = [w for w in words if w.lower() not in STOP_WORDS]
    if len(meaningful_words) == 0:
        return True

    # Is just a person's name pattern (1-3 capitalized words, no other content)
    name_pattern = re.compile(r"^[A-Z][a-z]+(\s+[A-Z][a-z]+){0,2}$")
    if name_pattern.match(title.strip()):
        return True

    return False


def extract_keywords(text: str, max_keywords: int = 8) -> list[str]:
    """Extract the most relevant keywords from text."""
    if not text:
        return []

    # Strip HTML and URLs before processing
    text = strip_html(text)
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"www\.\S+", " ", text)
    text = redact_pii(text)

    # First, check for known product mentions
    found_products = []
    for pattern, product_name in PRODUCT_PATTERNS:
        if pattern.search(text):
            found_products.append(product_name)

    # Tokenize and count meaningful words
    words = re.findall(r"[a-zA-Z]{3,}", text)
    # Filter: stop words, redaction artifacts, likely names, HTML remnants
    meaningful = []
    for w in words:
        w_lower = w.lower()
        if w_lower in STOP_WORDS:
            continue
        if len(w_lower) <= 2:
            continue
        if "redacted" in w_lower:
            continue
        if is_likely_name(w):
            continue
        meaningful.append(w_lower)

    word_counts = Counter(meaningful)

    # Get top keywords (excluding product names already found)
    product_words = set()
    for p in found_products:
        product_words.update(w.lower() for w in p.split())

    top_words = [
        word for word, _ in word_counts.most_common(max_keywords + 10)
        if word not in product_words
    ][:max_keywords]

    return found_products + top_words


# ---------------------------------------------------------------------------
# Title grammar & capitalization normalization
# ---------------------------------------------------------------------------

# Words that should stay lowercase in title case (unless first word)
_TITLE_CASE_LOWERCASE = {
    "a", "an", "the", "and", "but", "or", "nor", "for", "yet", "so",
    "in", "on", "at", "to", "by", "of", "as", "if",
    "vs", "via", "per", "from", "into", "with", "over", "than",
}

# Words/acronyms that should always be uppercase
_ALWAYS_UPPER = {
    "api", "dns", "ssl", "tls", "vpn", "sso", "mfa", "2fa", "pii",
    "sla", "cms", "csv", "url", "ip", "it", "ai", "qa", "ui", "ux",
    "aws", "gcp", "ccsk", "ccak", "cczt", "ccm", "caiq", "gdpr",
    "seo", "dmarc", "dkim", "spf", "mx", "csa", "pdf", "html",
    "saml", "oauth", "ldap", "smtp", "imap", "http", "https",
    "sftp", "ftp", "ssh", "sql", "json", "xml", "yaml", "rsa",
    "soc", "iso", "nist", "cis", "iam", "rbac", "cidr", "cdn",
    "grc", "404",
    # CSA-specific
    "taise", "orbs", "star", "caiq",
    # Common IT/business acronyms missing from original list
    "hr", "lms", "crm", "erp", "itsm", "itil",
}

# Brand names that have specific capitalization
_BRAND_CASING = {
    "zendesk": "Zendesk", "github": "GitHub", "gitlab": "GitLab",
    "airtable": "Airtable", "zapier": "Zapier", "slack": "Slack",
    "salesforce": "Salesforce", "pardot": "Pardot", "hubspot": "HubSpot",
    "jira": "Jira", "confluence": "Confluence", "zoom": "Zoom",
    "chatgpt": "ChatGPT", "claude": "Claude", "copilot": "Copilot",
    "cloudflare": "Cloudflare", "digitalocean": "DigitalOcean",
    "surveymonkey": "SurveyMonkey", "mailgun": "Mailgun",
    "google": "Google", "microsoft": "Microsoft", "wordpress": "WordPress",
    "skilljar": "Skilljar", "tableau": "Tableau", "okta": "Okta",
    "drupal": "Drupal", "mailchimp": "Mailchimp", "docusign": "DocuSign",
    "starwatch": "STARWatch", "linkedin": "LinkedIn",
    "rippling": "Rippling", "bamboohr": "BambooHR", "skilljar": "Skilljar",
    "pearsonvue": "Pearson VUE", "pearson": "Pearson",
}

# Common grammar fixes: (pattern, replacement)
_GRAMMAR_FIXES = [
    # Double spaces
    (re.compile(r"\s{2,}"), " "),
    # Missing space after punctuation (but not in URLs or acronyms)
    (re.compile(r"([a-z])\.([A-Z])"), r"\1. \2"),
    # Lowercase "i" standing alone
    (re.compile(r"\bi\b(?!\.\w)"), "I"),
    # "i'm", "i've", "i'd", "i'll" → proper case
    (re.compile(r"\bi'(m|ve|d|ll)\b", re.IGNORECASE), lambda m: f"I'{m.group(1).lower()}"),
    # "dont" → "don't", "cant" → "can't", "wont" → "won't", "doesnt" → "doesn't"
    (re.compile(r"\bdon'?t\b", re.IGNORECASE), "don't"),
    (re.compile(r"\bdont\b", re.IGNORECASE), "don't"),
    (re.compile(r"\bcan'?t\b", re.IGNORECASE), "can't"),
    (re.compile(r"\bcant\b", re.IGNORECASE), "can't"),
    (re.compile(r"\bwon'?t\b", re.IGNORECASE), "won't"),
    (re.compile(r"\bwont\b", re.IGNORECASE), "won't"),
    (re.compile(r"\bdoesn'?t\b", re.IGNORECASE), "doesn't"),
    (re.compile(r"\bdoesnt\b", re.IGNORECASE), "doesn't"),
    (re.compile(r"\bisn'?t\b", re.IGNORECASE), "isn't"),
    (re.compile(r"\bisnt\b", re.IGNORECASE), "isn't"),
    (re.compile(r"\bhasn'?t\b", re.IGNORECASE), "hasn't"),
    (re.compile(r"\bhasnt\b", re.IGNORECASE), "hasn't"),
    (re.compile(r"\bhaven'?t\b", re.IGNORECASE), "haven't"),
    (re.compile(r"\bhavent\b", re.IGNORECASE), "haven't"),
    (re.compile(r"\bwasn'?t\b", re.IGNORECASE), "wasn't"),
    (re.compile(r"\bwasnt\b", re.IGNORECASE), "wasn't"),
    (re.compile(r"\baren'?t\b", re.IGNORECASE), "aren't"),
    (re.compile(r"\barent\b", re.IGNORECASE), "aren't"),
    (re.compile(r"\bshouldn'?t\b", re.IGNORECASE), "shouldn't"),
    (re.compile(r"\bshouldnt\b", re.IGNORECASE), "shouldn't"),
    (re.compile(r"\bwouldn'?t\b", re.IGNORECASE), "wouldn't"),
    (re.compile(r"\bwouldnt\b", re.IGNORECASE), "wouldn't"),
    (re.compile(r"\bcouldn'?t\b", re.IGNORECASE), "couldn't"),
    (re.compile(r"\bcouldnt\b", re.IGNORECASE), "couldn't"),
    (re.compile(r"\bneed'?nt\b", re.IGNORECASE), "needn't"),
    # "im" → "I'm" (only standalone)
    (re.compile(r"\bim\b"), "I'm"),
    # Strip trailing whitespace/punctuation artifacts
    (re.compile(r"\s+$"), ""),
]


def _title_case_word(word: str, is_first: bool) -> str:
    """Apply proper title case to a single word."""
    lower = word.lower()

    # Check brand names first
    if lower in _BRAND_CASING:
        return _BRAND_CASING[lower]

    # Check acronyms
    if lower in _ALWAYS_UPPER:
        return word.upper()

    # Category prefix like [Tooling] — leave as-is
    if word.startswith("[") and word.endswith("]"):
        return word

    # Lowercase articles/prepositions unless first word
    if not is_first and lower in _TITLE_CASE_LOWERCASE:
        return lower

    # Capitalize first letter, preserve rest (handles "ChatGPT" → "ChatGPT")
    if word and word[0].islower():
        return word[0].upper() + word[1:]

    return word


def normalize_title_grammar(title: str) -> str:
    """Apply proper grammar, capitalization, and formatting to a suggested title.

    - Title case (with smart exceptions for articles, prepositions, acronyms, brands)
    - Fix common contractions and grammar issues
    - Preserve category prefix brackets [Category]
    - Handle special casing for known product/brand names
    """
    if not title or not title.strip():
        return title

    # Extract category prefix if present, process the rest separately
    prefix = ""
    body = title
    bracket_match = re.match(r"^(\[.+?\])\s*", title)
    if bracket_match:
        prefix = bracket_match.group(1) + " "
        body = title[bracket_match.end():]

    if not body.strip():
        return title

    # Apply grammar fixes first (on the body)
    for pattern, replacement in _GRAMMAR_FIXES:
        body = pattern.sub(replacement, body)

    # Fix multi-word brand names before title-casing individual words
    _MULTI_WORD_BRANDS = [
        (re.compile(r"\bdigital\s*ocean\b", re.IGNORECASE), "DigitalOcean"),
        (re.compile(r"\bgoogle\s+drive\b", re.IGNORECASE), "Google Drive"),
        (re.compile(r"\bgoogle\s+sheet[s]?\b", re.IGNORECASE), "Google Sheets"),
        (re.compile(r"\bgoogle\s+doc[s]?\b", re.IGNORECASE), "Google Docs"),
    ]
    for pat, replacement in _MULTI_WORD_BRANDS:
        body = pat.sub(replacement, body)

    # Preserve filenames (e.g., security.txt, robots.txt) — mark them to skip title-casing
    _filename_re = re.compile(r"\b(\w+\.(?:txt|json|xml|yaml|yml|csv|html|css|js|py|sh|md|log|cfg|conf|ini|env))\b", re.IGNORECASE)
    _file_placeholders = {}
    def _protect_filename(m):
        placeholder = f"__FILE{len(_file_placeholders)}__"
        _file_placeholders[placeholder] = m.group(0)
        return placeholder
    body = _filename_re.sub(_protect_filename, body)

    # Apply title case word by word
    words = body.split()
    result_words = []
    for i, word in enumerate(words):
        is_first = (i == 0)

        # Handle words with internal punctuation (e.g., "email—needs", "dns/domain")
        # Split on em-dash and slash, title-case each part
        if "—" in word and word != "—":
            parts = word.split("—")
            parts = [_title_case_word(p, is_first or j == 0) for j, p in enumerate(parts) if p]
            result_words.append("—".join(parts))
        elif word in ("—", "–", "-"):
            result_words.append(word)
        elif "/" in word and not word.startswith("http"):
            parts = word.split("/")
            parts = [_title_case_word(p, is_first) for p in parts if p]
            result_words.append("/".join(parts))
        else:
            result_words.append(_title_case_word(word, is_first))

    body = " ".join(result_words)

    # Restore protected filenames
    for placeholder, original in _file_placeholders.items():
        body = body.replace(placeholder, original)

    # Ensure body starts with uppercase after all processing
    if body and body[0].islower():
        body = body[0].upper() + body[1:]

    return f"{prefix}{body}".strip()


def build_suggested_title(title: str, description: str, comments: list[dict]) -> str:
    """Build a suggested title from the ticket description and comments."""
    # Strip URLs from title so they don't pollute keyword extraction
    title_for_keywords = re.sub(r"https?://\S+", " ", title)
    title_for_keywords = re.sub(r"www\.\S+", " ", title_for_keywords).strip()

    # Combine text sources
    all_text = description or ""
    for c in comments[:3]:
        body = c.get("plain_body") or c.get("body", "")
        if body:
            all_text += " " + body

    all_text = strip_html(all_text[:3000])

    # Extract keywords
    keywords = extract_keywords(all_text)

    if not keywords:
        return ""

    # Check for common action patterns in the description
    desc_clean = strip_html((description or "")[:1500]).lower()
    action = ""

    action_patterns = [
        # Token/voucher issues (very common for CSA)
        (r"(?:have not|haven'?t|did not|didn'?t|never)\s+(?:received?|got|get)\s+(?:(?:my|the|an?)\s+)?(token|voucher|exam|certificate|badge|receipt|confirmation|access|login)",
         "Not received: {match}"),
        # Where/how to find things
        (r"(?:where|how)\s+(?:can|do)\s+i\s+(?:find|get|download|access|view|see|locate|retrieve)\s+(?:(?:my|the|an?)\s+)?(token|voucher|exam|certificate|receipt|course|training|badge)",
         "Question: How to find {match}"),
        # Purchase followed by issue
        (r"i\s+(?:purchased|bought|paid\s+for|ordered)\s+(?:(?:the|a|an)\s+)?(.+?)(?:\s+(?:but|and|however|yet)\s+)",
         "Purchased — {match}"),
        # Can't find specific things
        (r"(?:can'?t|cannot|unable\s+to|couldn'?t)\s+(?:find|locate|see|view)\s+(?:(?:my|the|an?)\s+)?(token|voucher|exam|certificate|receipt|course|account|login)",
         "Cannot find {match}"),
        # Credential/account status issues
        (r"(?:my|the)\s+(certificate|exam|token|voucher|badge|account|login|password|access|membership)\s+(?:is|has|was)\s+(expired?|missing|invalid|wrong|incorrect|locked|blocked|suspended)",
         "{match} issue"),
        # Receipt/refund requests
        (r"(?:need|want|would\s+like|requesting?)\s+(?:a\s+)?(receipt|invoice|refund|credit|reimbursement|proof\s+of\s+purchase)",
         "Request: {match}"),
        # Original patterns
        (r"(can't|cannot|unable to|couldn't|could not)\s+(access|log\s*in|sign\s*in|connect|view|open|download|upload|use|find|see|load|reach)",
         "Cannot {match}"),
        (r"(need|want|would like|requesting|request)\s+(?:to\s+)?(add|remove|change|update|reset|create|delete|modify|enable|disable|upgrade|renew|transfer|consolidate)",
         "Request to {match}"),
        (r"(how|where)\s+(?:do|can|to)\s+(\w+(?:\s+\w+){0,3})",
         "Question: How to {match}"),
        (r"(error|failed|failure|crash|broken|bug|not working|doesn'?t work|isn'?t working)",
         "Error"),
        (r"(expired?|expir(?:ing|ation))",
         "Expiration issue"),
        (r"(invoice|billing|charge|payment|refund|credit)",
         "Billing inquiry"),
        (r"(certificate|certification|exam|badge|credential)",
         "Certification inquiry"),
        (r"(registry|listing|profile|entry)",
         "Registry/listing inquiry"),
    ]

    _articles = {"a", "an", "the", "my", "your", "our", "their"}

    for pattern, template in action_patterns:
        match = re.search(pattern, desc_clean)
        if match:
            if "{match}" in template:
                groups = [g for g in match.groups() if g and g.strip() not in _articles]
                relevant = groups[-1] if groups else ""
                relevant = relevant.strip()
                clean_words = [w for w in relevant.split()
                               if w.lower() not in STOP_WORDS
                               and "redacted" not in w.lower()
                               and not is_likely_name(w.capitalize())]
                if clean_words:
                    action = template.replace("{match}", " ".join(clean_words))
                else:
                    action = template.split("{")[0].strip()
            else:
                action = template
            break

    # Build the title
    products = [k for k in keywords if any(p == k for _, p in PRODUCT_PATTERNS)]
    other_keywords = [k for k in keywords if k not in products
                      and "redacted" not in k.lower()
                      and not is_likely_name(k.capitalize())][:4]

    if products and action:
        suggested = f"{products[0]}: {action}"
    elif products:
        context = " ".join(other_keywords[:3]).capitalize() if other_keywords else "inquiry"
        suggested = f"{products[0]} — {context}"
    elif action:
        context = " ".join(other_keywords[:3]) if other_keywords else ""
        if context:
            suggested = f"{action} — {context}"
        else:
            suggested = action
    else:
        # Fallback: title-case each keyword rather than raw lowercase concatenation
        titled = " ".join(w.capitalize() for w in other_keywords[:5])
        suggested = titled if titled else ""

    # Final cleanup
    suggested = suggested.strip(" —-:")
    suggested = re.sub(r"\s+", " ", suggested)
    # Remove any remaining redaction markers
    suggested = re.sub(r"\[?\w*_?REDACTED\]?", "", suggested, flags=re.IGNORECASE).strip()
    suggested = re.sub(r"\s+", " ", suggested).strip(" —-:")

    # Capitalize first letter
    if suggested:
        suggested = suggested[0].upper() + suggested[1:]

    # Check for incomplete suggestions (Issues 3 & 4)
    # Reject if too short, too generic, or ends with incomplete phrase
    if len(suggested) < 15:
        return ""
    
    incomplete_endings = ("to", "how to", "how", "what", "where", "for", "about", "with", "cannot")
    if any(suggested.lower().endswith(ending) for ending in incomplete_endings):
        return ""
    
    # Check for broken/generic patterns like "Cannot" alone or "Question: How to" without object
    if suggested.lower() in ("cannot", "error", "question", "error: ") \
            or (suggested.lower().startswith(("question: how to", "cannot ")) and len(suggested) < 25):
        return ""

    # Check for multiple emails in body indicating list/table (Issue 5)
    if description:
        email_count = len(re.findall(r"\w+@\w+", description))
        title_word_count = len(title.split())
        # If many emails but vague/short title, skip suggestion
        if email_count >= 3 and title_word_count <= 3:
            return ""

    # Prefix with category
    category = detect_category(title, description)
    suggested = f"[{category}] {suggested}"

    # Enforce max length (align with MAX_TITLE_LENGTH constant)
    if len(suggested) > MAX_TITLE_LENGTH:
        suggested = suggested[:MAX_TITLE_LENGTH - 3] + "..."

    return suggested if len(suggested) >= 10 else ""


# Phrases that are meta-commentary about ticket creation, not the actual request
_META_COMMENTARY_PATTERNS = [
    re.compile(r"(?:I\s+)?(?:never\s+)?(?:did|didn'?t)\s+find\s+a\s+ticket", re.IGNORECASE),
    re.compile(r"(?:did\s+we|have\s+we)\s+creat\w*\s+a\s+ticket", re.IGNORECASE),
    re.compile(r"(?:I\s+am|I'?m)\s+creat(?:ing|e)\s+(?:a\s+)?(?:this|one|ticket|request)", re.IGNORECASE),
    re.compile(r"(?:opening|submitting|filing|raising)\s+(?:this|a)\s+(?:ticket|request|issue)", re.IGNORECASE),
    # NOTE: "as discussed" is handled as a strippable prefix, not a full skip
    re.compile(r"(?:hi|hello|hey|good\s+(?:morning|afternoon|evening))\b[,.]?\s*(?:team|all|everyone|there)?", re.IGNORECASE),
    re.compile(r"hope\s+(?:this|you|all\s+is)", re.IGNORECASE),
    re.compile(r"please\s+(?:see|find)\s+(?:below|attached|the\s+(?:below|attached))", re.IGNORECASE),
    re.compile(r"(?:I\s+)?(?:want|would\s+like)\s+to\s+(?:report|flag|bring\s+to)", re.IGNORECASE),
    re.compile(r"^(?:hi|hello|hey|dear)\s+", re.IGNORECASE),
    re.compile(r"following\s+up\s+on", re.IGNORECASE),
    re.compile(r"(?:this\s+is|here\s+is)\s+(?:a\s+)?(?:follow.?up|reminder|request)", re.IGNORECASE),
    re.compile(r"thank\s+you\s+(?:for|in\s+advance)", re.IGNORECASE),
    re.compile(r"^(?:please\s+)?(?:can\s+you|could\s+you)\s+(?:please\s+)?(?:help|assist)\b", re.IGNORECASE),
    re.compile(r"^apologi(?:es|ze)", re.IGNORECASE),
    re.compile(r"(?:could\s+not|couldn'?t|cannot|can'?t)\s+find\s+a?\s*(?:related|existing)?\s*ticket", re.IGNORECASE),
    re.compile(r"I\s+was\s+just\s+looking\s+at", re.IGNORECASE),
]


def _is_meta_sentence(sentence: str) -> bool:
    """Check if a sentence is meta-commentary rather than the actual request."""
    for pat in _META_COMMENTARY_PATTERNS:
        if pat.search(sentence):
            return True
    return False


def _build_best_effort_title(category: str, title: str, description: str, comments: list[dict]) -> str:
    """Build a best-effort title for tickets that failed normal suggestion.

    Uses a relaxed approach: extract the most meaningful phrases from the
    description and combine with the detected category.  The result is
    intended as a *starting point* for human review, not a final answer.
    """
    desc = (description or "")[:3000]
    # Strip HTML but preserve newlines for signature detection
    desc = re.sub(r"<br\s*/?>", "\n", desc, flags=re.IGNORECASE)
    desc = re.sub(r"<[^>]+>", " ", desc)
    desc = re.sub(r"&nbsp;?", " ", desc, flags=re.IGNORECASE)
    desc = re.sub(r"&amp;?", " and ", desc, flags=re.IGNORECASE)
    desc = re.sub(r"&#?\w+;", " ", desc)
    # Strip URLs
    desc = re.sub(r"https?://\S+", " ", desc)
    desc = re.sub(r"www\.\S+", " ", desc)
    # Strip email attribution lines ("On <date> <name> wrote:")
    desc = re.sub(r"On\s+\w{3},\s+\w{3}\s+\d{1,2},\s+\d{4}\s+at\s+.{0,80}wrote:", "\n", desc, flags=re.IGNORECASE)
    # Strip email headers and forwarded message markers
    desc = re.sub(r"^\s*-+\s*Forwarded message\s*-+\s*$", "\n", desc, flags=re.MULTILINE | re.IGNORECASE)
    desc = re.sub(r"^\s*(?:From|To|Subject|Date|Cc|Bcc):\s+.*$", "\n", desc, flags=re.MULTILINE)
    # Strip signature blocks: lines with job titles, org names, contact info
    desc = re.sub(r"^\s*(?:Technology|Vice\s+President|Director|Manager|Content\s+Development|Senior|Lead|Chief|Head\s+of)\b.*$", "\n", desc, flags=re.MULTILINE)
    desc = re.sub(r"^\s*Cloud Security Alliance\s*$", "\n", desc, flags=re.MULTILINE)
    desc = re.sub(r"^\s*[pe]:\s*(?:\+?\d[\d\s.()\-]+|\S+@\S+)\s*$", "\n", desc, flags=re.MULTILINE)
    desc = re.sub(r"^\s*[me]:\s*(?:\+?\d[\d\s.()\-]+|\S+@\S+)\s*$", "\n", desc, flags=re.MULTILINE)
    # Strip standalone name lines (2-3 capitalized words on own line, common in sigs)
    desc = re.sub(r"^\s*[A-Z][a-z]+\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?\s*$", "\n", desc, flags=re.MULTILINE)
    # Strip signature dividers and boilerplate
    desc = re.sub(r"[-–—]{2,}.*", "\n", desc)
    desc = re.sub(r"(?:^|\n)\s*(?:Sent from|Submitted from|Get Outlook|Cheers|Best|Regards|Thanks)[\s,]*(?:\n|$)", "\n", desc, flags=re.IGNORECASE)

    # Strippable meta-prefixes: "As discussed last week, ..." → keep the rest
    _PREFIX_STRIP = re.compile(
        r"^(?:as\s+(?:discussed|mentioned|per\s+our\s+\w+)\s+(?:last\s+week|earlier|yesterday|today|on\s+\w+day)"
        r"[,;]?\s*(?:I\s+)?(?:wanted?\s+to\s+(?:chat\s+on|discuss|talk\s+about|ask\s+about)\s*)?)",
        re.IGNORECASE,
    )

    # Try to grab the first meaningful non-meta sentence
    sentences = re.split(r"[.\n!?]+", desc)
    best_sentence = ""
    for s in sentences:
        s = s.strip()
        if not s:
            continue
        # Skip meta-commentary (greetings, "I'm creating a ticket", etc.)
        if _is_meta_sentence(s):
            continue
        # Strip meta-prefixes but keep the actionable remainder
        s = _PREFIX_STRIP.sub("", s).strip()
        words = [w for w in s.split() if w.lower() not in STOP_WORDS and len(w) > 2
                 and "redacted" not in w.lower()]
        if len(words) >= 3:
            # Truncate to reasonable length
            best_sentence = s.strip()
            if len(best_sentence) > 70:
                best_sentence = best_sentence[:67].rsplit(" ", 1)[0] + "..."
            break

    if not best_sentence:
        # Fall back to top keywords
        keywords = extract_keywords(desc)
        if keywords:
            best_sentence = " ".join(keywords[:5]).capitalize()

    if not best_sentence:
        return ""

    # Clean up: capitalize first letter, strip trailing punctuation artifacts
    best_sentence = best_sentence[0].upper() + best_sentence[1:] if best_sentence else ""
    best_sentence = best_sentence.rstrip(" ,;:-–—")

    suggested = f"[{category}] {best_sentence}"
    if len(suggested) > 100:
        suggested = suggested[:97] + "..."
    return suggested


# ---------------------------------------------------------------------------
# Title enhancement — clean, fix, and enrich descriptive titles
# ---------------------------------------------------------------------------

# Common IT/ops typos: misspelling → correction
_TYPO_FIXES = [
    (re.compile(r"\bassesments?\b", re.IGNORECASE), lambda m: "assessments" if m.group().endswith("s") else "assessment"),
    (re.compile(r"\breplacment\b", re.IGNORECASE), "replacement"),
    (re.compile(r"\brecieve\b", re.IGNORECASE), "receive"),
    (re.compile(r"\brecieved\b", re.IGNORECASE), "received"),
    (re.compile(r"\boccured\b", re.IGNORECASE), "occurred"),
    (re.compile(r"\boccurr?ance\b", re.IGNORECASE), "occurrence"),
    (re.compile(r"\bseperates?\b", re.IGNORECASE), lambda m: "separates" if m.group().endswith("s") else "separate"),
    (re.compile(r"\bneccessary\b", re.IGNORECASE), "necessary"),
    (re.compile(r"\bneccessity\b", re.IGNORECASE), "necessity"),
    (re.compile(r"\bdependan(cy|cies)\b", re.IGNORECASE), lambda m: "dependency" if m.group(1) == "cy" else "dependencies"),
    (re.compile(r"\bprivelege\b", re.IGNORECASE), "privilege"),
    (re.compile(r"\bpriviledge\b", re.IGNORECASE), "privilege"),
    (re.compile(r"\benviroment\b", re.IGNORECASE), "environment"),
    (re.compile(r"\bmanagment\b", re.IGNORECASE), "management"),
    (re.compile(r"\bmaintainance\b", re.IGNORECASE), "maintenance"),
    (re.compile(r"\bmaintenence\b", re.IGNORECASE), "maintenance"),
    (re.compile(r"\binfrastucture\b", re.IGNORECASE), "infrastructure"),
    (re.compile(r"\bauthentification\b", re.IGNORECASE), "authentication"),
    (re.compile(r"\bsubcription\b", re.IGNORECASE), "subscription"),
    (re.compile(r"\bconfigration\b", re.IGNORECASE), "configuration"),
    (re.compile(r"\bdecommision\b", re.IGNORECASE), "decommission"),
    (re.compile(r"\bdecomission\b", re.IGNORECASE), "decommission"),
    (re.compile(r"\bWIndows\b"), "Windows"),
]

# Noise patterns to strip from titles
_TITLE_NOISE = [
    # [AAS.Tasks.ScheduledTask], [AAS.IO.InternalOperationsRequest] etc.
    re.compile(r"\s*[-–—]\s*\[AAS\.[^\]]+\]", re.IGNORECASE),
    re.compile(r"\s*\[AAS\.[^\]]+\]", re.IGNORECASE),
    # Calendar/notification noise: "@ Fri Aug 29, 2025 12:00 (MDT) (CSA IT)"
    re.compile(r"\s*@\s*\w{3}\s+\w{3}\s+\d{1,2},?\s*\d{4}\s+\d{1,2}:\d{2}\s*(?:\([^)]+\)\s*)*", re.IGNORECASE),
    # "Notification:" prefix
    re.compile(r"^Notification:\s*", re.IGNORECASE),
    # Trailing URLs
    re.compile(r"\s+https?://\S+\s*$"),
    # Trailing question marks on requests (not actual questions)
    re.compile(r"\?+\s*$"),
    # Trailing periods and spaces
    re.compile(r"[.\s]+$"),
    # " - Cloud Security Alliance" suffix
    re.compile(r"\s*[-–—]\s*Cloud Security Alliance\s*$", re.IGNORECASE),
]

# Conversational/informal patterns to clean up
_INFORMAL_PATTERNS = [
    # "Can you help with X?" → "X"
    (re.compile(r"^(?:can\s+you\s+(?:help|assist)\s+(?:with|me\s+with)\s*)", re.IGNORECASE), ""),
    # "Please ensure X" → "Ensure X"
    (re.compile(r"^please\s+", re.IGNORECASE), ""),
    # "Regarding the X" → "X"
    (re.compile(r"^regarding\s+(?:the\s+)?", re.IGNORECASE), ""),
    # "Proposal: X — 15 min to walk through?" → "Proposal: X"
    (re.compile(r"\s*[-–—]\s*\d+\s*min(?:ute)?s?\s+to\s+(?:walk\s+through|discuss|chat|review)\s*\??$", re.IGNORECASE), ""),
    # "Important X update:" → "X Update:"
    (re.compile(r"^important\s+", re.IGNORECASE), ""),
]


def enhance_title(title: str, description: str, comments: list[dict]) -> str:
    """Enhance a descriptive title using content from description and comments.

    Cleans noise, fixes typos, truncates overly long titles, and supplements
    short titles with context extracted from the ticket body.
    Returns the enhanced title WITHOUT the [Category] prefix.
    """
    enhanced = title.strip()

    # --- Phase 1: Strip noise ---
    for noise_pat in _TITLE_NOISE:
        enhanced = noise_pat.sub("", enhanced).strip()

    # --- Phase 2: Fix typos ---
    for typo_pat, fix in _TYPO_FIXES:
        enhanced = typo_pat.sub(fix, enhanced)

    # --- Phase 3: Clean up informal/conversational phrasing ---
    for informal_pat, repl in _INFORMAL_PATTERNS:
        enhanced = informal_pat.sub(repl, enhanced).strip()

    # --- Phase 4: Handle overly long titles (>110 chars) ---
    # Try to extract a tighter summary from description
    if len(enhanced) > 110:
        desc = strip_html((description or "")[:2000])
        desc_clean = re.sub(r"https?://\S+", " ", desc)
        desc_clean = re.sub(r"\s+", " ", desc_clean).strip()

        # Meta-commentary phrases that should never become titles
        _meta_reject = re.compile(
            r"(?:fine.?tune|tweak|adjust|look\s+(?:at|into)|figure\s+out|"
            r"check\s+(?:on|this|that)|circle\s+back|follow\s+up|take\s+a\s+look|"
            r"approve\s+(?:the\s+)?access\??|confirm\s+(?:that|this)|"
            r"let\s+(?:me|us)\s+know|thanks\s*[!?]?\s*$|"
            r"^\s*(?:this|that|it)\s+(?:a\s+bit|out|up))",
            re.IGNORECASE,
        )

        # Build set of meaningful words from the original title for relevance check
        _title_words = {w.lower() for w in re.findall(r"[a-zA-Z]{3,}", enhanced)}
        _title_words -= STOP_WORDS

        # Look for a clear action/request sentence in the description
        # that might be more concise than the title
        _request_patterns = [
            re.compile(r"(?:we\s+need\s+to|need\s+to|please|request(?:ing)?)\s+(.{15,70}?)(?:\.|$)", re.IGNORECASE),
            re.compile(r"(?:goal|objective|purpose)(?:\s+is)?(?:\s*:\s*|\s+to\s+)(.{15,70}?)(?:\.|$)", re.IGNORECASE),
        ]
        for rp in _request_patterns:
            m = rp.search(desc_clean[:500])
            if m:
                candidate = m.group(1).strip().rstrip(" ,;:-–—?!")
                # Reject meta-commentary candidates
                if _meta_reject.search(candidate):
                    continue
                # Require topical overlap: at least 1 meaningful word in common with title
                _cand_words = {w.lower() for w in re.findall(r"[a-zA-Z]{3,}", candidate)}
                _cand_words -= STOP_WORDS
                if not (_cand_words & _title_words):
                    continue
                # Only use if meaningfully shorter and still informative
                if len(candidate) < len(enhanced) - 15 and len(candidate.split()) >= 4:
                    enhanced = candidate
                    break

        # If still too long, truncate at a natural break point
        if len(enhanced) > 110:
            # Try to cut at a natural boundary (dash, comma, period)
            for sep in [" — ", " – ", " - ", ", ", ": "]:
                idx = enhanced.find(sep, 40)
                if 40 < idx < 100:
                    enhanced = enhanced[:idx]
                    break
            else:
                # Hard truncate at word boundary
                if len(enhanced) > 110:
                    enhanced = enhanced[:107].rsplit(" ", 1)[0] + "..."

    # --- Phase 5: Supplement short titles with description context ---
    # If the title is descriptive but very short (<25 chars), add context
    word_count = len(enhanced.split())
    if word_count <= 4 and len(enhanced) < 30 and description:
        desc = strip_html((description or "")[:1500])
        desc = re.sub(r"https?://\S+", " ", desc)
        desc = re.sub(r"\s+", " ", desc).strip().lower()

        # Extract a short context phrase from the description
        _context_patterns = [
            # "I would like to X" / "We need to X"
            re.compile(r"(?:i\s+would\s+like\s+to|we\s+need\s+to|i\s+want\s+to|need\s+to|plan\s+to)\s+(.{10,50}?)(?:\.|,|\n|$)", re.IGNORECASE),
            # "for X" / "to X" purpose phrases
            re.compile(r"(?:this\s+is\s+(?:for|about|to))\s+(.{10,40}?)(?:\.|,|\n|$)", re.IGNORECASE),
            # "set up X for Y"
            re.compile(r"(?:set\s*up|configure|create|build)\s+(.{10,40}?)(?:\.|,|\n|$)", re.IGNORECASE),
        ]
        for cp in _context_patterns:
            m = cp.search(desc[:500])
            if m:
                context = m.group(1).strip().rstrip(" ,;:-–—")
                context_words = [w for w in context.split() if w.lower() not in STOP_WORDS and len(w) > 2]
                if len(context_words) >= 2:
                    # Append as a dash-separated clarification
                    supplement = context[0].upper() + context[1:]
                    if len(supplement) > 40:
                        supplement = supplement[:37].rsplit(" ", 1)[0] + "..."
                    enhanced = f"{enhanced} — {supplement}"
                    break

    # Final cleanup
    enhanced = enhanced.strip()
    enhanced = re.sub(r"\s{2,}", " ", enhanced)
    if enhanced and enhanced[0].islower():
        enhanced = enhanced[0].upper() + enhanced[1:]

    return enhanced


def suggest_title(ticket: dict, comments: list[dict]) -> dict:
    """Analyze a ticket title and suggest improvements (with grammar normalization)."""
    result = _suggest_title_raw(ticket, comments)

    # Normalize title casing on all suggestions for consistency
    suggested = result.get("suggested_title", "")
    if suggested:
        # Preserve the [Category] prefix, normalize the rest
        prefix_match = re.match(r"^(\[[^\]]+\]\s*)", suggested)
        if prefix_match:
            prefix = prefix_match.group(1)
            body = suggested[len(prefix):]
            result["suggested_title"] = prefix + normalize_title_grammar(body)
        else:
            result["suggested_title"] = normalize_title_grammar(suggested)

    return result


def _build_detailed_reason(original: str, cleaned: str, enhanced: str, category: str) -> str:
    """Compare original vs. final title and list every change made."""
    changes: list[str] = []

    # 1. Re:/Fwd: stripped
    if re.match(r"^(?:(?:re|fw|fwd)\s*:\s*)+", original, re.IGNORECASE):
        changes.append("Stripped Re:/Fwd: prefix")

    # 2. System tags removed
    if re.search(r"\[AAS\.[^\]]+\]", original, re.IGNORECASE):
        changes.append("Removed system tag (e.g. [AAS...])")

    # 3. Org suffix removed
    if re.search(r"[-–—]\s*Cloud Security Alliance\s*$", original, re.IGNORECASE):
        changes.append("Removed org suffix")

    if re.search(r"\[CloudSecurityAlliance\]", original, re.IGNORECASE):
        changes.append("Removed [CloudSecurityAlliance] tag")

    # 4. Capitalization changes (title case)
    if cleaned and enhanced and cleaned.lower() == enhanced.lower() and cleaned != enhanced:
        changes.append("Fixed capitalization (title case)")
    elif cleaned and enhanced and cleaned.lower() != enhanced.lower():
        # Content was actually rewritten/enhanced
        changes.append("Enhanced wording")

    # 5. Grammar fixes — check for specific patterns
    if cleaned != enhanced:
        # Check for common grammar fix signals
        if re.search(r"\bi\b", cleaned) and not re.search(r"\bI\b", cleaned):
            changes.append("Grammar: capitalized 'I'")
        # Contraction fixes
        for wrong, _right in [("dont", "don't"), ("cant", "can't"), ("wont", "won't"),
                               ("isnt", "isn't"), ("doesnt", "doesn't"), ("its a", "it's a")]:
            if wrong in cleaned.lower() and wrong not in enhanced.lower():
                changes.append(f"Grammar: fixed '{wrong}'")
                break

    # 6. Category prefix added
    changes.append(f"Added [{category}] prefix")

    if not changes:
        return f"Adding [{category}] prefix for triage"

    return "; ".join(changes)


def _suggest_title_raw(ticket: dict, comments: list[dict]) -> dict:
    """Analyze a ticket title using heuristic rules and suggest improvements."""
    current_title = ticket.get("subject", ticket.get("raw_subject", ""))
    description = ticket.get("description", "")

    # Clean subject for analysis (preserve original for report)
    cleaned_title = clean_subject_line(current_title)

    # Check for PII in title FIRST
    if title_contains_pii(current_title):
        return {
            "suggested_title": redact_pii(current_title),
            "status": "PII in Title",
            "reason": "Current title contains personal information (email/phone) that should be removed",
        }

    # Check for spam/marketing tickets
    if is_spam_ticket(cleaned_title, description):
        return {
            "suggested_title": "",
            "status": "Likely Spam",
            "reason": "Title matches spam/marketing patterns — likely not a real support ticket",
        }

    # Check tickets whose title already has a [Category] prefix
    existing_prefix_match = re.match(r"^\[(.+?)\]\s", current_title)
    if existing_prefix_match:
        existing_category = existing_prefix_match.group(1)
        # If the existing category is valid and matches what we'd detect, skip
        if existing_category in VALID_CATEGORIES:
            title_body = current_title[existing_prefix_match.end():]
            correct_category = detect_category(title_body, description)
            if correct_category == existing_category:
                return {
                    "suggested_title": "",
                    "status": "Already Categorized",
                    "reason": "Title already has a correct category prefix — skipping",
                }
            # Category is wrong — suggest a correction
            corrected = f"[{correct_category}] {title_body}"
            return {
                "suggested_title": corrected,
                "status": "Suggestion",
                "reason": f"Recategorized from [{existing_category}] to [{correct_category}]",
            }
        # Unknown/custom prefix (e.g. [OPS-PROJ | P## | T#/#]) — always skip
        return {
            "suggested_title": "",
            "status": "Already Categorized",
            "reason": "Title already has a category prefix — skipping",
        }

    # Check for automation/notification tickets
    if is_automation_ticket(cleaned_title, description):
        # Special case: STAR Submission contact form titles
        # Format: "STAR Submission <ID> from <email>" — strip ID, keep domain
        _star_submission_re = re.compile(
            r"^(?:invoice\s+requested\s*~\s*)?STAR\s+(?:Contact\s+Form|Submission)\s+\S+\s+from\s+(\S+@(\S+))",
            re.IGNORECASE,
        )
        _star_m = _star_submission_re.match(cleaned_title)
        if _star_m:
            domain = _star_m.group(2).rstrip(".")
            suggested = f"[Needs Triage] New STAR Submission — {domain}"
            return {
                "suggested_title": suggested,
                "status": "Suggestion",
                "reason": "STAR submission contact form — stripped form ID, preserved submitter domain",
            }

        # Try to extract product name from multiple automation title formats
        _automation_extractors = [
            (re.compile(r"purchase notification for (.+)", re.IGNORECASE), "Purchase Notification"),
            (re.compile(r"registration notification for (.+)", re.IGNORECASE), "Registration"),
            (re.compile(r"(?:fwd?:\s*)?your purchase of (.+)", re.IGNORECASE), "Purchase Notification"),
            (re.compile(r"receipt for (?:your )?(?:purchase of )?(.+)", re.IGNORECASE), "Receipt"),
            (re.compile(r"order confirmation[:\s]+(.+)", re.IGNORECASE), "Order Confirmation"),
            (re.compile(r"payment of \$[\d.,]+ from .+ for (.+)", re.IGNORECASE), "Payment Received"),
            (re.compile(r"valid.?ai.?ted\b", re.IGNORECASE), "Valid-AI-ted Notification"),
        ]
        for pat, notif_type in _automation_extractors:
            m = pat.search(current_title)
            if m:
                product_name = m.group(1).strip()
                product_name = re.sub(r"\s*[-–—]\s*Cloud Security Alliance\s*$", "", product_name).rstrip(".")
                if len(product_name) > 80:
                    product_name = product_name[:77] + "..."
                category = detect_category(current_title, description)
                suggested = f"[{category}] {product_name} — {notif_type}"
                return {
                    "suggested_title": suggested,
                    "status": "Suggestion",
                    "reason": "Title is from an automated/notification email — needs a human-readable subject",
                }

        # Try to build a better title from the content
        suggested = build_suggested_title(cleaned_title, description, comments)
        reason = classify_vagueness_or_automation(cleaned_title, description)
        if suggested and suggested.lower() != cleaned_title.lower():
            validated = validate_suggestion(suggested, ticket["id"])
            if validated:
                return {
                    "suggested_title": validated,
                    "status": "Suggestion",
                    "reason": reason,
                }

        # Fallback: build a best-effort title from description/comments
        category = detect_category(cleaned_title, description)
        best_effort = _build_best_effort_title(category, cleaned_title, description, comments)
        if best_effort:
            return {
                "suggested_title": best_effort,
                "status": "Suggestion",
                "reason": "Title is from an automated/notification email; Added [" + category + "] prefix",
            }
        return {
            "suggested_title": "",
            "status": "Needs Manual Review",
            "reason": "Title is from an automated/notification email — needs a human-readable subject",
        }

    if not is_vague_title(cleaned_title):
        # Title is descriptive — enhance it and add category prefix
        if not re.match(r"^\[.+?\]", cleaned_title):
            enhanced = enhance_title(cleaned_title, description, comments)
            category = detect_category(enhanced, description)
            prefixed = f"[{category}] {enhanced}"
            if len(prefixed) > 130:
                prefix_part = f"[{category}] "
                max_enhanced = 130 - len(prefix_part)
                truncated = enhanced[:max_enhanced].rsplit(" ", 1)[0]
                prefixed = f"{prefix_part}{truncated}..."
            # Build a detailed reason listing each change made
            reason = _build_detailed_reason(current_title, cleaned_title, enhanced, category)
            return {
                "suggested_title": prefixed,
                "status": "Suggestion",
                "reason": reason,
            }
        return {
            "suggested_title": "",
            "status": "Keep Current",
            "reason": "Title is already descriptive and categorized",
        }

    # Title is vague — try to build a better one
    suggested = build_suggested_title(cleaned_title, description, comments)

    if suggested and suggested.lower() != cleaned_title.lower():
        # Validate the suggestion
        validated = validate_suggestion(suggested, ticket["id"])
        if validated:
            reason = classify_vagueness_or_automation(cleaned_title, description)
            return {
                "suggested_title": validated,
                "status": "Suggestion",
                "reason": reason,
            }

    # Generate a best-effort suggestion for manual review
    category = detect_category(cleaned_title, description)

    # For URL titles, try to derive a readable title from the URL path first
    if is_url_title(current_title):
        url_title = _title_from_url(current_title)
        if url_title:
            prefixed = f"[{category}] {url_title}"
            validated = validate_suggestion(prefixed, ticket["id"])
            if validated:
                return {
                    "suggested_title": validated,
                    "status": "Suggestion",
                    "reason": "Title was a URL — replaced with descriptive subject; Added [" + category + "] prefix",
                }

    best_effort = _build_best_effort_title(category, cleaned_title, description, comments)

    if is_url_title(current_title):
        reason = "Title is a URL — suggested title is a best-effort guess and should be reviewed"
    else:
        reason = "Title is vague — suggested title is a best-effort guess and should be reviewed"
    return {
        "suggested_title": best_effort,
        "status": "Needs Manual Review",
        "reason": reason,
    }


def classify_vagueness_or_automation(title: str, description: str = "") -> str:
    """Return a human-readable reason why the title was flagged."""
    # Check automation first
    if is_automation_ticket(title, description):
        for pattern in AUTOMATION_PATTERNS:
            if pattern.search(title.strip()):
                return "Title is from an automated/notification email — needs a human-readable subject"
        return "Ticket body indicates this was auto-generated — title may not describe the actual issue"

    if is_url_title(title):
        return "Title is a URL — needs a human-readable subject describing the actual request"

    cleaned = title.strip().lower()
    cleaned = re.sub(r"^(re|fw|fwd)\s*:\s*", "", cleaned).strip()

    if cleaned in {"", "none", "n/a", "na", "no subject", "(no subject)", "...", ".", "-", "--", "___"}:
        return "Title is empty or placeholder"

    words = cleaned.split()
    if len(words) == 1:
        return f"Single-word title \"{title.strip()}\" — too vague for triage"

    if len(cleaned) < 10:
        return f"Title too short ({len(cleaned)} chars) to identify the issue"

    meaningful = [w for w in words if w not in STOP_WORDS]
    if len(meaningful) == 0:
        return "Title contains only generic/filler words"

    name_pattern = re.compile(r"^[A-Z][a-z]+(\s+[A-Z][a-z]+){0,2}$")
    if name_pattern.match(title.strip()):
        return "Title appears to be a person's name, not a description"

    if cleaned in VAGUE_TITLES:
        return f"Generic title \"{title.strip()}\" — doesn't describe the specific issue"

    return "Title lacks specificity for effective triage"


def validate_suggestion(suggestion: str, ticket_id: int) -> str | None:
    if not suggestion or not suggestion.strip():
        logger.warning("Ticket #%s: Empty suggestion received, skipping.", ticket_id)
        return None

    suggestion = suggestion.strip().strip('"').strip("'")

    if len(suggestion) > MAX_TITLE_LENGTH:
        logger.warning(
            "Ticket #%s: Suggestion too long (%d chars), skipping: %s",
            ticket_id, len(suggestion), suggestion[:80] + "...",
        )
        return None

    pii_leak_patterns = [
        re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"),
        re.compile(r"\b\d{3}-\d{2}-\d{4}\b"),
        re.compile(r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b"),
    ]
    for pattern in pii_leak_patterns:
        if pattern.search(suggestion):
            logger.warning(
                "Ticket #%s: Suggestion appears to contain PII, skipping.", ticket_id,
            )
            return None

    # Also reject if redaction markers leaked through
    if re.search(r"redacted", suggestion, re.IGNORECASE):
        logger.warning("Ticket #%s: Suggestion contains redaction markers, skipping.", ticket_id)
        return None

    return suggestion


# ---------------------------------------------------------------------------
# Spreadsheet builder
# ---------------------------------------------------------------------------

DARK_HEADER   = "1F2D3D"
SUMMARY_BG    = "E8EAF6"
SUGGEST_FILL  = "E8F4E8";  ALT_SUGGEST  = "F0FAF0"
KEEP_FILL     = "F5F5F5";  ALT_KEEP     = "FAFAFA"
ERROR_FILL    = "FFE8E8";  ALT_ERROR    = "FFF0F0"
SKIP_FILL     = "FFF9C4";  ALT_SKIP     = "FFFDE7"
PII_FILL      = "FFF3E0";  ALT_PII      = "FFF8E8"
SPAM_FILL     = "EEEEEE";  ALT_SPAM     = "F5F5F5"
LINK_COLOR    = "1155CC"
SUGGEST_BADGE = "27AE60"
KEEP_BADGE    = "7F8C8D"
ERROR_BADGE   = "C0392B"
SKIP_BADGE    = "F57F17"
PII_BADGE     = "E65100"
SPAM_BADGE    = "9E9E9E"

HEADERS = [
    "Ticket #", "Title", "Action", "Reason",
    "Ticket Status", "Requester", "Last Updated",
]
WIDTHS = [10, 52, 18, 40, 12, 20, 13]


# ---------------------------------------------------------------------------
# Duplicate / related ticket detection
# ---------------------------------------------------------------------------

def _tokenize_for_similarity(text: str) -> set[str]:
    """Extract meaningful lowercase tokens from text for similarity comparison."""
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"\[.+?\]", " ", text)  # strip [Category] prefixes
    text = re.sub(r"[^\w\s]", " ", text)
    words = text.lower().split()
    return {w for w in words if w not in STOP_WORDS and len(w) > 2
            and "redacted" not in w}


def _jaccard_similarity(set_a: set, set_b: set) -> float:
    """Compute Jaccard similarity between two token sets."""
    if not set_a or not set_b:
        return 0.0
    intersection = set_a & set_b
    union = set_a | set_b
    return len(intersection) / len(union)


def detect_related_tickets(tickets_data: list[dict], threshold: float = 0.35) -> dict[int, list[tuple[int, float]]]:
    """Find related/duplicate tickets based on title + description keyword overlap.

    Args:
        tickets_data: list of dicts with keys: ticket_id, title, description, category
        threshold: minimum Jaccard similarity to consider tickets related (0.0–1.0)

    Returns:
        dict mapping ticket_id → list of (related_ticket_id, similarity_score)
        sorted by similarity descending.
    """
    # Pre-compute token sets for each ticket
    ticket_tokens = {}
    for t in tickets_data:
        tid = t["ticket_id"]
        # Combine title (weighted by repeating) + first 500 chars of description
        title_text = t.get("title", "")
        desc_text = (t.get("description", "") or "")[:500]
        # Title tokens are more important — include them twice
        combined = f"{title_text} {title_text} {desc_text}"
        ticket_tokens[tid] = {
            "tokens": _tokenize_for_similarity(combined),
            "category": t.get("category", ""),
            "title": title_text,
        }

    # Pairwise comparison
    tids = list(ticket_tokens.keys())
    related: dict[int, list[tuple[int, float]]] = {tid: [] for tid in tids}

    for i in range(len(tids)):
        for j in range(i + 1, len(tids)):
            tid_a, tid_b = tids[i], tids[j]
            data_a, data_b = ticket_tokens[tid_a], ticket_tokens[tid_b]

            sim = _jaccard_similarity(data_a["tokens"], data_b["tokens"])

            # Boost similarity if they share the same category
            if data_a["category"] and data_a["category"] == data_b["category"]:
                sim *= 1.15  # 15% boost for same category

            # Boost if titles share 2+ meaningful words
            title_tokens_a = _tokenize_for_similarity(data_a["title"])
            title_tokens_b = _tokenize_for_similarity(data_b["title"])
            title_overlap = title_tokens_a & title_tokens_b
            if len(title_overlap) >= 2:
                sim *= 1.25  # 25% boost for title keyword overlap

            if sim >= threshold:
                related[tid_a].append((tid_b, round(sim, 2)))
                related[tid_b].append((tid_a, round(sim, 2)))

    # Sort each ticket's related list by similarity descending
    for tid in related:
        related[tid].sort(key=lambda x: x[1], reverse=True)

    return related


def format_related_tickets(related: list[tuple[int, float]]) -> str:
    """Format related ticket list for display in spreadsheet cell."""
    if not related:
        return ""
    parts = []
    for tid, score in related[:3]:  # max 3 related tickets per cell
        pct = int(score * 100)
        parts.append(f"#{tid} ({pct}%)")
    return ", ".join(parts)


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, fc="000000",
          bg=None, wrap=False, align="left", size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fc, size=size)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = _border()
    return c


def _status_colors(status):
    if status == "Suggestion":
        return (SUGGEST_FILL, ALT_SUGGEST, SUGGEST_BADGE)
    elif status == "Error":
        return (ERROR_FILL, ALT_ERROR, ERROR_BADGE)
    elif status == "Skipped":
        return (SKIP_FILL, ALT_SKIP, SKIP_BADGE)
    elif status == "PII in Title":
        return (PII_FILL, ALT_PII, PII_BADGE)
    elif status == "Needs Manual Review":
        return (SKIP_FILL, ALT_SKIP, SKIP_BADGE)
    elif status == "Likely Spam":
        return (SPAM_FILL, ALT_SPAM, SPAM_BADGE)
    return (KEEP_FILL, ALT_KEEP, KEEP_BADGE)


def format_date(iso_string: str | None) -> str:
    if not iso_string:
        return ""
    try:
        dt = datetime.fromisoformat(iso_string.replace("Z", "+00:00"))
        return dt.strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return str(iso_string)[:10]


def write_xlsx_report(rows: list[dict], output_path: str, run_meta: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Title Suggestions"

    # ── Summary rows ──────────────────────────────────────────────────────
    summary_items = [
        ("Tickets Scanned:",  str(run_meta["tickets_scanned"]),  "1F2D3D"),
        ("Suggestions Made:", str(run_meta["suggestions_made"]), SUGGEST_BADGE),
        ("Needs Review:",     str(run_meta.get("manual_reviews", 0)), "F57F17"),
        ("Titles Kept:",      str(run_meta["titles_kept"]),      KEEP_BADGE),
        ("Errors:",           str(run_meta["errors"]),           ERROR_BADGE),
    ]
    for sr, (label, val, color) in enumerate(summary_items, 1):
        c = ws.cell(row=sr, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=12, color=color)
        c.fill = PatternFill("solid", start_color=SUMMARY_BG)
        c.alignment = Alignment(horizontal="right", vertical="center")

        c2 = ws.cell(row=sr, column=2, value=int(val))
        c2.font = Font(name="Arial", bold=True, size=14, color=color)
        c2.fill = PatternFill("solid", start_color=SUMMARY_BG)
        c2.alignment = Alignment(horizontal="left", vertical="center")

        for col in range(3, len(HEADERS) + 1):
            sc = ws.cell(row=sr, column=col)
            sc.fill = PatternFill("solid", start_color=SUMMARY_BG)

    ws.row_dimensions[len(summary_items) + 1].height = 6  # spacer

    # ── Header row ────────────────────────────────────────────────────────
    HEADER_ROW = len(summary_items) + 2
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color=DARK_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HEADER_ROW].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────
    cur_row = HEADER_ROW + 1
    for r in rows:
        status = r.get("Status", "Keep Current")
        fill_main, fill_alt, badge_color = _status_colors(status)
        even = cur_row % 2 == 0
        bg = fill_main if not even else fill_alt

        # Col 1: Ticket # (hyperlinked)
        ticket_id = r.get("Ticket #", "")
        url = r.get("Ticket URL", "")
        lnk = ws.cell(row=cur_row, column=1, value=ticket_id)
        lnk.font = Font(name="Arial", bold=True, color=LINK_COLOR, underline="single", size=11)
        lnk.alignment = Alignment(horizontal="center", vertical="top")
        if url:
            lnk.hyperlink = url
        lnk.fill = PatternFill("solid", start_color=bg)
        lnk.border = _border()

        # Col 2: Title (current + suggested combined)
        current = r.get("Current Title", "")
        suggested = r.get("Suggested Title", "")
        if status == "Suggestion" and suggested:
            # Rich text: current title on first line, arrow + suggested on second
            title_cell = ws.cell(row=cur_row, column=2)
            title_cell.value = f"{current}\n-> {suggested}"
            title_cell.font = Font(name="Arial", size=11, color=badge_color)
            title_cell.fill = PatternFill("solid", start_color=bg)
            title_cell.alignment = Alignment(wrap_text=True, vertical="top")
            title_cell.border = _border()
        else:
            _cell(ws, cur_row, 2, current, bg=bg, wrap=True)

        # Col 3: Action (merged Status + Recommendation)
        _cell(ws, cur_row, 3, r.get("Recommendation", ""), bold=True, fc=badge_color, bg=bg, align="center")

        # Col 4: Reason
        _cell(ws, cur_row, 4, r.get("Reason", ""), bg=bg, wrap=True, size=10)

        # Col 5: Ticket Status
        _cell(ws, cur_row, 5, r.get("Ticket Status", ""), bg=bg, align="center")

        # Col 6: Requester
        _cell(ws, cur_row, 6, r.get("Requester", ""), bg=bg, align="center")

        # Col 7: Last Updated
        _cell(ws, cur_row, 7, r.get("Last Updated", ""), bg=bg, align="center")

        ws.row_dimensions[cur_row].height = 48
        cur_row += 1

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── Claude Prompts sheet ────────────────────────────────────────────
    # One prompt per actionable suggestion so Claude can update titles via
    # the Zendesk API (or MCP tool).
    actionable = [r for r in rows
                  if r.get("Status") in ("Suggestion", "PII in Title")
                  and r.get("Suggested Title")]
    if actionable:
        cp = wb.create_sheet("Claude Prompts")

        # Header row
        prompt_headers = ["Ticket #", "Claude Prompt"]
        prompt_widths = [10, 110]
        for ci, (h, w) in enumerate(zip(prompt_headers, prompt_widths), 1):
            c = cp.cell(row=1, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", start_color=DARK_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
            cp.column_dimensions[get_column_letter(ci)].width = w
        cp.row_dimensions[1].height = 24

        pr = 2  # prompt row counter
        for r in actionable:
            tid = r.get("Ticket #", "")
            current = r.get("Current Title", "")
            suggested = r.get("Suggested Title", "")

            prompt = (
                f"Update the subject/title of Zendesk ticket #{tid} "
                f"from its current title \"{current}\" "
                f"to the new title \"{suggested}\". "
                f"Use the Zendesk API: PUT /api/v2/tickets/{tid} "
                f"with payload: {{\"ticket\": {{\"subject\": \"{suggested}\"}}}}."
            )

            even = pr % 2 == 0
            bg = "F7F9FC" if even else "FFFFFF"

            # Col 1: Ticket #
            tid_cell = cp.cell(row=pr, column=1, value=tid)
            tid_cell.font = Font(name="Arial", bold=True, color=LINK_COLOR, underline="single", size=11)
            tid_cell.alignment = Alignment(horizontal="center", vertical="top")
            ticket_url = r.get("Ticket URL", "")
            if ticket_url:
                tid_cell.hyperlink = ticket_url
            tid_cell.fill = PatternFill("solid", start_color=bg)
            tid_cell.border = _border()

            # Col 2: Claude Prompt
            _cell(cp, pr, 2, prompt, bg=bg, wrap=True, size=10)

            cp.row_dimensions[pr].height = 48
            pr += 1

        # Freeze header
        cp.freeze_panes = "A2"

    # ── Category Taxonomy sheet ───────────────────────────────────────────
    ct = wb.create_sheet("Category Taxonomy")
    tax_headers = ["Prefix", "Description"]
    for ci, (h, w) in enumerate(zip(tax_headers, [28, 60]), 1):
        c = ct.cell(row=1, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color=DARK_HEADER)
        c.alignment = Alignment(horizontal="left", vertical="top")
        c.border = _border()
        ct.column_dimensions[get_column_letter(ci)].width = w
    ct.row_dimensions[1].height = 24

    for ri, (prefix, desc) in enumerate(sorted(VALID_CATEGORIES.items()), 2):
        even = ri % 2 == 0
        bg = "F7F9FC" if even else "FFFFFF"
        _cell(ct, ri, 1, f"[{prefix}]", bold=True, bg=bg)
        _cell(ct, ri, 2, desc, bg=bg, wrap=True)

    wb.save(output_path)
    logger.info("Spreadsheet saved → %s (%d data rows)", output_path, len(rows))


# -- Google Drive upload -----------------------------------------------------

def upload_to_gdrive(file_path):
    if not GDRIVE_AVAILABLE:
        print("  [Drive] google-auth libraries not installed -- skipping upload.")
        return
    if not GDRIVE_SA_JSON or not GDRIVE_FOLDER_ID:
        print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON or GDRIVE_FOLDER_ID not set -- skipping.")
        return

    try:
        sa_json = GDRIVE_SA_JSON.strip()
        if not sa_json:
            print("  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON is blank after stripping whitespace -- skipping.")
            return
        creds_info = json.loads(sa_json)

        if creds_info.get("type") == "service_account":
            creds = service_account.Credentials.from_service_account_info(
                creds_info,
                scopes=["https://www.googleapis.com/auth/drive"],
            )
        else:
            from google.oauth2.credentials import Credentials
            creds = Credentials(
                token=creds_info.get("token"),
                refresh_token=creds_info["refresh_token"],
                token_uri=creds_info.get("token_uri", "https://oauth2.googleapis.com/token"),
                client_id=creds_info["client_id"],
                client_secret=creds_info["client_secret"],
                scopes=creds_info.get("scopes"),
            )

        service = build("drive", "v3", credentials=creds)

        file_name = os.path.basename(file_path)
        file_metadata = {"name": file_name, "parents": [GDRIVE_FOLDER_ID]}
        media = MediaFileUpload(
            file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )

        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, name, webViewLink",
            supportsAllDrives=True,
        ).execute()

        print(f"  [Drive] Uploaded: {uploaded['name']}")
        print(f"  [Drive] View at : {uploaded.get('webViewLink', '(no link)')}")

    except json.JSONDecodeError as e:
        print(f"  [Drive] GDRIVE_SERVICE_ACCOUNT_JSON is not valid JSON: {e}")
        print(f"  [Drive] Secret starts with: {repr(GDRIVE_SA_JSON[:80])}")
        print("  [Drive] Check that the secret contains the raw JSON (not base64 or a file path).")
    except Exception as e:
        print(f"  [Drive] Upload failed: {e}")


def main():
    missing = []
    for var in ("ZENDESK_SUBDOMAIN", "ZENDESK_EMAIL", "ZENDESK_API_TOKEN"):
        if not os.environ.get(var):
            missing.append(var)
    if missing:
        logger.error("Missing required environment variables: %s", ", ".join(missing))
        sys.exit(1)

    logger.info("=" * 60)
    logger.info("Zendesk Ticket Title Suggester (Rule-Based)")
    logger.info("=" * 60)
    logger.info("Mode: LOG ONLY (no tickets will be modified)")
    logger.info("Engine: Heuristic rules (no AI API required)")
    logger.info("Max tickets: %d", MAX_TICKETS)
    logger.info("Rate limits: Zendesk=%.1fs", ZENDESK_RATE_LIMIT_DELAY)
    logger.info("Retry config: max_retries=%d, base_delay=%.1fs", MAX_RETRIES, RETRY_BASE_DELAY)
    logger.info("PII redaction: ENABLED")
    logger.info("=" * 60)

    logger.info("Fetching open tickets from Zendesk (%s)...", ZENDESK_SUBDOMAIN)
    try:
        tickets = fetch_open_tickets()
    except requests.RequestException as e:
        logger.error("Failed to fetch tickets from Zendesk after retries: %s", e)
        sys.exit(1)

    logger.info("Found %d open tickets to analyze.", len(tickets))

    # Batch-fetch requester names so every row shows a real name
    all_requester_ids = [t.get("requester_id") for t in tickets if t.get("requester_id")]
    if all_requester_ids:
        logger.info("Fetching requester names for %d users...", len(set(all_requester_ids)))
        fetch_user_names(list(set(all_requester_ids)))

    report_rows: list[dict] = []
    tickets_for_similarity: list[dict] = []  # for duplicate detection
    suggestion_count = 0
    keep_count = 0
    manual_review_count = 0
    spam_count = 0
    errors = 0

    for i, ticket in enumerate(tickets, 1):
        ticket_id = ticket["id"]
        current_title = ticket.get("subject", ticket.get("raw_subject", ""))
        ticket_status = ticket.get("status", "")
        created_at = ticket.get("created_at", "")
        updated_at = ticket.get("updated_at", "")
        ticket_url = f"https://{ZENDESK_SUBDOMAIN}.zendesk.com/agent/tickets/{ticket_id}"

        # Resolve requester name from cache (batch-fetched above)
        requester_id = ticket.get("requester_id")
        requester_name = get_user_name(requester_id) if requester_id else ""

        logger.info("[%d/%d] Analyzing ticket #%s: %s", i, len(tickets), ticket_id, current_title)

        try:
            comments = fetch_ticket_comments(ticket_id)
        except requests.RequestException as e:
            logger.error("  → Failed to fetch comments for ticket #%s: %s", ticket_id, e)
            errors += 1
            report_rows.append({
                "Ticket #": ticket_id,
                "Status": "Error",
                "Current Title": current_title,
                "Suggested Title": "",
                "Recommendation": "Review Manually",
                "Reason": f"Failed to fetch comments: {str(e)[:80]}",
                "Ticket URL": ticket_url,
                "Ticket Status": ticket_status.capitalize(),
                "Requester": requester_name,
                "Created": format_date(created_at),
                "Last Updated": format_date(updated_at),
            })
            continue

        result = suggest_title(ticket, comments)

        suggested_title = result["suggested_title"]
        status = result["status"]
        reason = result["reason"]

        if status == "Suggestion":
            suggestion_count += 1
            recommendation = "Update Title"
            logger.info("  → Suggested: %s", suggested_title if suggested_title else "(flag only — no auto-suggestion)")
        elif status == "Needs Manual Review":
            manual_review_count += 1
            recommendation = "Review Manually"
            logger.info("  → Needs manual review: vague title with no auto-suggestion")
        elif status == "Likely Spam":
            spam_count += 1
            recommendation = "Likely Spam"
            logger.info("  → Likely spam/marketing ticket")
        elif status == "Error":
            errors += 1
            recommendation = "Review Manually"
            logger.info("  → Error analyzing title.")
        else:
            keep_count += 1
            recommendation = "No Action Needed"
            logger.info("  → Title is fine, no change suggested.")

        description = ticket.get("description", "")
        category = detect_category(clean_subject_line(current_title), description)

        report_rows.append({
            "Ticket #": ticket_id,
            "Status": status,
            "Current Title": current_title,
            "Suggested Title": suggested_title,
            "Recommendation": recommendation,
            "Reason": reason,
            "Related Tickets": "",  # populated after duplicate detection
            "Ticket URL": ticket_url,
            "Ticket Status": ticket_status.capitalize(),
            "Requester": requester_name,
            "Created": format_date(created_at),
            "Last Updated": format_date(updated_at),
        })

        tickets_for_similarity.append({
            "ticket_id": ticket_id,
            "title": clean_subject_line(current_title),
            "description": description,
            "category": category,
        })

    # ── Duplicate / related ticket detection ──────────────────────────────
    logger.info("Running duplicate/related ticket detection...")
    related_map = detect_related_tickets(tickets_for_similarity)
    related_groups_count = 0
    for row in report_rows:
        tid = row["Ticket #"]
        related_list = related_map.get(tid, [])
        if related_list:
            row["Related Tickets"] = format_related_tickets(related_list)
            related_groups_count += 1
    logger.info("Found %d tickets with related/duplicate matches.", related_groups_count)

    # Print summary
    print("\n" + "=" * 80)
    print(f"TITLE SUGGESTION REPORT — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)
    print(f"Tickets scanned: {len(tickets)}")
    print(f"Suggestions made: {suggestion_count}")
    print(f"Titles kept: {keep_count}")
    print(f"Manual reviews needed: {manual_review_count}")
    print(f"Spam/marketing flagged: {spam_count}")
    print(f"Errors encountered: {errors}")
    print(f"Engine: Rule-based heuristics (no AI API)")
    print(f"PII redaction: enabled")
    print("=" * 80)

    for row in report_rows:
        if row["Status"] == "Suggestion":
            print(f"\nTicket #{row['Ticket #']}  {row['Ticket URL']}")
            print(f"  Current:   {row['Current Title']}")
            print(f"  Suggested: {row['Suggested Title'] or '(review needed — no auto-suggestion)'}")
            print(f"  Reason:    {row['Reason']}")

    print("\n" + "=" * 80)

    run_meta = {
        "run_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tickets_scanned": len(tickets),
        "suggestions_made": suggestion_count,
        "titles_kept": keep_count,
        "manual_reviews": manual_review_count,
        "spam_flagged": spam_count,
        "related_tickets": related_groups_count,
        "errors": errors,
    }

    write_xlsx_report(report_rows, REPORT_PATH, run_meta)

    upload_to_gdrive(REPORT_PATH)

    if suggestion_count == 0 and manual_review_count == 0:
        logger.info("All ticket titles look good — nothing to suggest!")

    if errors > 0 and errors == len(tickets):
        logger.error("All tickets failed to process. Exiting with error.")
        sys.exit(1)


if __name__ == "__main__":
    main()
